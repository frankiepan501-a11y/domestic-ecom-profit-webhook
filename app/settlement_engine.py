"""Settlement-file-driven domestic e-commerce gross-profit engine.

This engine keeps the old order-driven engine as fallback, but the finance
confirmation path must use platform settlement files as the ledger:
Tmall transaction settlement, Douyin settlement orders, XHS product settlement,
and JD arrival/no-settlement evidence.
"""
from __future__ import annotations

import csv
import hashlib
import io
import math
import re
from collections import defaultdict
from datetime import date, datetime
from typing import Any

import openpyxl
import xlrd


COMPANY = "深圳奥迪尔"

MONTHLY_HEADER = [
    "月份", "平台", "店铺", "主体", "销售订单数", "销量", "销售额(RMB)", "退款(RMB)",
    "净销售额(RMB)", "平台费用(RMB)", "广告费(RMB)", "采购成本(RMB)", "尾程费用(RMB)",
    "其他费用(RMB)", "试算毛利(RMB)", "试算毛利率", "结算回款/净回款(RMB)",
    "P0缺口数", "P1注意项", "状态",
]
PRODUCT_HEADER = [
    "国内电商平台名称", "运营人员", "国家", "站点", "月份", "MSKU", "中文名称", "销量",
    "退款数量", "销售额(RMB)", "退款(RMB)", "平台服务费(RMB)", "广告费(RMB)",
    "采购成本(RMB)", "尾程费用(RMB)", "其他成本(RMB)", "毛利润(RMB)", "毛利率", "备注",
]
COST_HEADER = [
    "平台", "店铺", "月份", "订单号", "ERP_SKU", "品名", "净成本数量", "单件采购成本",
    "采购成本", "成本来源",
]
LOG_HEADER = [
    "平台", "店铺", "月份", "订单号", "子订单号", "运单号", "承运商", "分摊尾程费用",
    "账单文件/API", "账单sheet", "来源", "状态",
]
FEE_HEADER = [
    "平台", "店铺", "月份", "来源文件", "费用类别", "订单号", "金额", "取数字段", "备注",
]
GAP_HEADER = ["P级", "平台", "店铺", "月份", "类别", "对象", "问题", "影响", "建议动作"]
SOURCE_HEADER = ["平台", "店铺", "资料类型", "文件名", "记录数/匹配数", "状态", "备注"]
NOTE_HEADER = ["主题", "说明"]
TAX_HEADER = [
    "平台", "店铺", "月份", "资料文件", "资料类型", "记录数", "读取状态", "差异/备注",
    "财务确认建议",
]


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)) and float(value).is_integer():
            return str(int(value))
    return str(value).strip().lstrip("\ufeff").strip("'").replace("\t", "").strip()


def money(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return 0.0
        return float(value)
    s = norm(value).replace(",", "").replace("￥", "").replace("¥", "").replace("元", "")
    if s in ("", "-", "--", "暂无数据", "无退款申请"):
        return 0.0
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def mny(value: Any) -> float:
    return round(float(value or 0), 2)


def pct(numerator: float, denominator: float) -> float:
    return round(numerator / denominator, 6) if denominator else 0.0


def p(row: dict[str, Any] | None, name: str) -> Any:
    if not row:
        return ""
    target = norm(name)
    for key, value in row.items():
        if norm(key) == target:
            return value
    return ""


def row_obj(headers: list[Any], values: list[Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    seen: dict[str, int] = {}
    for i, raw_name in enumerate(headers):
        name = norm(raw_name) or f"空列{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        out[name] = values[i] if i < len(values) else ""
    return out


def read_csv(buf: bytes, prefer_gbk: bool = False) -> list[dict[str, Any]]:
    encodings = ("gb18030", "gbk", "utf-8-sig", "utf-8") if prefer_gbk else (
        "utf-8-sig", "utf-8", "gb18030", "gbk"
    )
    rows: list[list[str]] = []
    for enc in encodings:
        try:
            text = buf.decode(enc)
            rows = list(csv.reader(io.StringIO(text)))
            break
        except Exception:
            continue
    if not rows:
        return []
    headers = rows[0]
    out: list[dict[str, Any]] = []
    for values in rows[1:]:
        if "".join(norm(v) for v in values) == "":
            continue
        out.append(row_obj(headers, values))
    return out


def sheet_rows(buf: bytes, filename: str, sheet_name: str = "", header_row: int = 1,
               key_col: int | None = None) -> list[dict[str, Any]]:
    if not buf:
        return []
    rows: list[list[Any]] = []
    head = buf[:8]
    if head[:4] == b"PK\x03\x04":
        def _read_xlsx(read_only: bool) -> list[list[Any]]:
            wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True, read_only=read_only)
            try:
                ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
                return [list(row) for row in ws.iter_rows(values_only=True)]
            finally:
                wb.close()

        rows = _read_xlsx(read_only=True)
        # Some Taobao/SF exports have a wrong worksheet dimension; read_only mode
        # then returns only the title row. Reopen normal mode only for that case.
        if len(rows) <= header_row:
            rows = _read_xlsx(read_only=False)
    elif head[:4] == b"\xd0\xcf\x11\xe0":
        wb = xlrd.open_workbook(file_contents=buf)
        if sheet_name and sheet_name in wb.sheet_names():
            sh = wb.sheet_by_name(sheet_name)
        else:
            sh = wb.sheet_by_index(0)
        for r in range(sh.nrows):
            values: list[Any] = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        values.append(xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode))
                    except Exception:
                        values.append(cell.value)
                else:
                    values.append(cell.value)
            rows.append(values)
    else:
        return read_csv(buf, prefer_gbk="天猫" in filename or "交易货款" in filename)

    if len(rows) < header_row:
        return []
    headers = rows[header_row - 1]
    out: list[dict[str, Any]] = []
    for values in rows[header_row:]:
        if not values or "".join(norm(v) for v in values) == "":
            continue
        if key_col is not None:
            key = values[key_col - 1] if key_col - 1 < len(values) else ""
            if norm(key) == "":
                continue
        out.append(row_obj(headers, values))
    return out


def canonical_shop(platform: str, shop: str, filename: str = "") -> str:
    text = f"{platform} {shop} {filename}".lower()
    if "天猫" in text or "tmall" in text:
        if "纷岚" in text or "funlab" in text or "梵乐璞" in text:
            return "天猫纷岚"
        if "powkong" in text or "宝空" in text or "宝宝" in text:
            return "天猫宝空"
    if "抖音" in text or "douyin" in text:
        if "纷岚" in text or "funlab" in text or "梵乐璞" in text:
            return "抖音纷岚"
        if "powkong" in text or "宝空" in text or "宝宝" in text:
            return "抖音宝空"
    if "小红书" in text or "xhs" in text:
        if "纷岚" in text or "funlab" in text or "梵乐璞" in text:
            return "小红书纷岚"
        if "powkong" in text or "宝空" in text or "宝宝" in text:
            return "小红书宝空"
    if "京东" in text or "jd" in text:
        if "纷岚" in text or "funlab" in text or "梵乐璞" in text:
            return "京东纷岚"
        if "powkong" in text or "宝空" in text or "宝宝" in text:
            return "京东宝空"
    return norm(shop) or norm(filename)


def platform_of_shop(shop: str) -> str:
    for platform in ("天猫", "抖音", "小红书", "京东"):
        if shop.startswith(platform):
            return platform
    return ""


def _fname(sf: dict[str, Any]) -> str:
    return norm(sf.get("fname") or sf.get("name") or "")


def _basename(path: str) -> str:
    return norm(path).replace("\\", "/").split("/")[-1]


def _files(raw: dict, platform: str | None = None, shop: str | None = None) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for sf in raw.get("source_files", []):
        fname = _fname(sf)
        sf_platform = norm(sf.get("platform"))
        sf_shop = canonical_shop(sf_platform, norm(sf.get("shop")), fname)
        if platform and sf_platform not in (platform, "全平台", ""):
            if platform not in fname:
                continue
        if shop and sf_shop != shop and shop not in fname:
            continue
        out.append(sf)
    return out


def _pick_file(files: list[dict[str, Any]], *keywords: str) -> dict[str, Any] | None:
    for sf in sorted(files, key=lambda x: _fname(x)):
        fname = _fname(sf)
        if fname.startswith("~$"):
            continue
        if all(k.lower() in fname.lower() for k in keywords):
            return sf
    return None


def _matching_files(files: list[dict[str, Any]], *keywords: str) -> list[dict[str, Any]]:
    out = []
    for sf in sorted(files, key=lambda x: _fname(x)):
        fname = _fname(sf)
        if fname.startswith("~$"):
            continue
        if all(k.lower() in fname.lower() for k in keywords):
            out.append(sf)
    return out


def _is_order_detail_export(filename: str) -> bool:
    """Recognize official order and item detail exports used to enrich settlement rows."""
    return any(keyword in filename for keyword in (
        "订单明细", "订单查询", "ExportOrderList", "ExportItemlList", "ExportItemList",
    ))


class SettlementReport:
    def __init__(self, year_month: str):
        self.year_month = year_month
        self.monthly: list[list[Any]] = []
        self.cost_rows: list[list[Any]] = []
        self.log_rows: list[list[Any]] = []
        self.fee_rows: list[list[Any]] = []
        self.gaps: list[list[Any]] = []
        self.source_rows: list[list[Any]] = []
        self.tax_rows: list[list[Any]] = []
        self.product: dict[tuple[str, str, str, str], dict[str, float | str]] = {}

    def add_gap(self, level: str, platform: str, shop: str, category: str, obj: str,
                problem: str, impact: str, action: str) -> None:
        self.gaps.append([level, platform, shop, self.year_month, category, obj, problem, impact, action])

    def gap_count(self, platform: str, shop: str, level: str) -> int:
        return sum(1 for g in self.gaps if g[0] == level and g[1] == platform and g[2] == shop)

    def add_source(self, platform: str, shop: str, typ: str, filename: str, count: Any,
                   status: str, note: str) -> None:
        self.source_rows.append([platform, shop, typ, filename, count, status, note])

    def add_tax(self, platform: str, shop: str, filename: str, typ: str, count: Any,
                status: str, note: str) -> None:
        self.tax_rows.append([
            platform, shop, self.year_month, filename, typ, count, status, note,
            "财务确认前复核税务A/B差异；P0仅输出资料读取审计",
        ])

    def add_fee(self, platform: str, shop: str, filename: str, category: str, order_id: str,
                amount: float, field: str, note: str) -> None:
        self.fee_rows.append([platform, shop, self.year_month, filename, category, order_id,
                              mny(amount), field, note])

    def add_cost(self, platform: str, shop: str, order_id: str, sku: str, name: str,
                 qty: float, unit: float, amount: float, source: str) -> None:
        self.cost_rows.append([platform, shop, self.year_month, order_id, sku, name,
                               mny(qty), mny(unit), mny(amount), source])

    def add_log(self, platform: str, shop: str, order_id: str, sub_id: str, waybill: str,
                carrier: str, amount: float, bill_file: str, bill_sheet: str, source: str,
                status: str) -> None:
        self.log_rows.append([platform, shop, self.year_month, order_id, sub_id, waybill, carrier,
                              mny(amount), bill_file, bill_sheet, source, status])

    def add_product(self, platform: str, shop: str, sku: str, name: str, qty: float,
                    refund_qty: float, sales: float, refund: float, platform_fee: float,
                    ad_fee: float, purchase: float, tail: float, other: float) -> None:
        key = (platform, shop, sku, name)
        if key not in self.product:
            self.product[key] = {
                "qty": 0.0, "refund_qty": 0.0, "sales": 0.0, "refund": 0.0,
                "platform_fee": 0.0, "ad_fee": 0.0, "purchase": 0.0, "tail": 0.0,
                "other": 0.0,
            }
        row = self.product[key]
        row["qty"] = float(row["qty"]) + qty
        row["refund_qty"] = float(row["refund_qty"]) + refund_qty
        row["sales"] = float(row["sales"]) + sales
        row["refund"] = float(row["refund"]) + refund
        row["platform_fee"] = float(row["platform_fee"]) + platform_fee
        row["ad_fee"] = float(row["ad_fee"]) + ad_fee
        row["purchase"] = float(row["purchase"]) + purchase
        row["tail"] = float(row["tail"]) + tail
        row["other"] = float(row["other"]) + other

    def add_monthly(self, platform: str, shop: str, orders: int, qty: float, sales: float,
                    refund: float, platform_fee: float, ad_fee: float, purchase: float,
                    tail: float, other: float, status: str) -> None:
        net = sales - refund
        profit = net - platform_fee - ad_fee - purchase - tail - other
        payback = net - platform_fee - ad_fee - other
        self.monthly.append([
            self.year_month, platform, shop, COMPANY, orders, mny(qty), mny(sales), mny(refund),
            mny(net), mny(platform_fee), mny(ad_fee), mny(purchase), mny(tail), mny(other),
            mny(profit), pct(profit, net), mny(payback),
            self.gap_count(platform, shop, "P0"), self.gap_count(platform, shop, "P1"), status,
        ])

    def product_rows(self) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for platform, shop, sku, name in sorted(self.product):
            r = self.product[(platform, shop, sku, name)]
            sales = float(r["sales"])
            refund = float(r["refund"])
            net = sales - refund
            profit = (
                net - float(r["platform_fee"]) - float(r["ad_fee"]) - float(r["purchase"])
                - float(r["tail"]) - float(r["other"])
            )
            rows.append([
                platform, "", "中国", shop, self.year_month, sku, name, mny(r["qty"]),
                mny(r["refund_qty"]), mny(sales), mny(refund), mny(r["platform_fee"]),
                mny(r["ad_fee"]), mny(r["purchase"]), mny(r["tail"]), mny(r["other"]),
                mny(profit), pct(profit, net),
                "平台/广告等不可直接归SKU费用按净销售额分摊；物流按运单拆分",
            ])
        return rows

    def notes(self) -> list[list[Any]]:
        return [
            ["试算定位", "本表是domestic-ecom-profit自动化试算，用于和人工基准逐字段A/B对账。"],
            ["资料范围", "当前P0按天猫、京东、抖音、小红书四个平台资料包；淘宝/拼多多暂缓。"],
            ["天猫归月", "天猫按交易货款账期/结算月口径；订单明细只补ERP_SKU和物流单号。"],
            ["广告费", "天猫广告取0947账户流水中交易日期=当月且交易类型=扣款或收支类型=支出的行；充值行排除。"],
            ["平台费用", "天猫平台费纳入主要费用CSV，排除-流水文件避免重复计入预冻结流水。"],
            ["抖音", "抖音按结算订单确认收入/退款/平台服务费/站外推广费；订单明细补SKU和快递。"],
            ["小红书", "小红书按商品结算明细确认入账/退款；退款数量回查订单查询包裹详情SKU件数扣减采购成本。"],
            ["物流", "按结算订单集合映射订单明细物流单号，再匹配前月/当月顺丰与中通账单；顺丰账单未命中时尝试EXP_RECE_QUERY_SFWAYBILL API。"],
            ["采购成本", "采购成本按ERP_SKU匹配产品采购成本台，优先采购成本(财务核算)，为空或0时用采购成本(ERP)，再兜底领星cg_price。"],
            ["京东", "京东按到账/结算明细和无结算确认输出零销售或费用行，不静默跳过。"],
        ]

    def as_dict(self) -> dict[str, Any]:
        return {
            "monthly_rows": self.monthly,
            "product_rows": self.product_rows(),
            "cost_rows": self.cost_rows,
            "log_rows": self.log_rows,
            "fee_rows": self.fee_rows,
            "gap_rows": self.gaps,
            "source_rows": self.source_rows,
            "note_rows": self.notes(),
            "tax_rows": self.tax_rows,
            "headers": {
                "月度毛利试算": MONTHLY_HEADER,
                "产品毛利_月度": PRODUCT_HEADER,
                "产品毛利_季度": PRODUCT_HEADER,
                "SKU成本明细": COST_HEADER,
                "物流匹配明细": LOG_HEADER,
                "费用明细汇总": FEE_HEADER,
                "缺口清单": GAP_HEADER,
                "资料清单审计": SOURCE_HEADER,
                "口径说明": NOTE_HEADER,
                "税务A_B核对": TAX_HEADER,
            },
        }


def cost_entry(cost_map: dict[str, dict[str, Any]], sku: str) -> tuple[float, str, str]:
    data = cost_map.get(norm(sku).upper()) or cost_map.get(norm(sku)) or {}
    return (
        float(data.get("unit_cost") or data.get("cost") or 0),
        norm(data.get("source")) or "成本缺失/为0",
        norm(data.get("name")),
    )


def build_bill_pool(raw: dict) -> dict[str, dict[str, Any]]:
    pool: dict[str, dict[str, Any]] = {}
    seen_bill_files: set[str] = set()
    for sf in raw.get("source_files", []):
        fname = _fname(sf)
        low = fname.lower()
        if fname.startswith("~$"):
            continue
        buf = sf.get("buf") or b""
        if ("顺丰" in fname or "中通" in fname) and fname.endswith((".xlsx", ".xls")):
            digest = hashlib.sha1(buf).hexdigest()
            if digest in seen_bill_files:
                continue
            seen_bill_files.add(digest)
        if "顺丰" in fname and fname.endswith((".xlsx", ".xls")):
            rows = sheet_rows(buf, fname, "账单明细", header_row=2, key_col=3)
            for r in rows:
                wb = norm(p(r, "运单号码"))
                if not wb or wb == "合计":
                    continue
                item = pool.setdefault(wb, {
                    "waybill": wb, "carrier": "顺丰", "amount": 0.0, "file": fname,
                    "sheet": "账单明细", "date": norm(p(r, "日期")), "source": "月结账单",
                })
                item["amount"] += money(p(r, "应付金额"))
        elif "中通" in fname and fname.endswith((".xlsx", ".xls")):
            rows = sheet_rows(buf, fname, "Sheet1", header_row=1, key_col=2)
            for r in rows:
                wb = norm(p(r, "运单号"))
                if not wb:
                    continue
                item = pool.setdefault(wb, {
                    "waybill": wb, "carrier": "中通", "amount": 0.0, "file": fname,
                    "sheet": "Sheet1", "date": norm(p(r, "账单日期")), "source": "月结账单",
                })
                item["amount"] += money(p(r, "合计"))
        elif "logistics" in low:
            continue
    for row in raw.get("logistics", []):
        wb = norm(row.get("tracking"))
        if not wb:
            continue
        source = norm(row.get("source")) or "parser"
        if "API" in source.upper() or "EXP_RECE" in source.upper() or row.get("api"):
            if wb not in pool:
                pool[wb] = {
                    "waybill": wb, "carrier": norm(row.get("carrier")) or "顺丰",
                    "amount": float(row.get("amount") or 0), "file": "顺丰API",
                    "sheet": "EXP_RECE_QUERY_SFWAYBILL", "date": norm(row.get("date")),
                    "source": "顺丰API",
                }
        elif wb not in pool:
            pool[wb] = {
                "waybill": wb, "carrier": norm(row.get("carrier")), "amount": float(row.get("amount") or 0),
                "file": norm(row.get("source")) or "parser", "sheet": "", "date": norm(row.get("date")),
                "source": norm(row.get("source")) or "parser",
            }
    return pool


def extract_sf_waybills(raw: dict, year_month: str) -> set[str]:
    out: set[str] = set()
    for sf in raw.get("source_files", []):
        fname = _fname(sf)
        if not fname.lower().endswith((".xlsx", ".xls", ".csv")):
            continue
        try:
            if _is_order_detail_export(fname):
                rows = sheet_rows(sf.get("buf") or b"", fname, "export") or sheet_rows(sf.get("buf") or b"", fname)
            else:
                continue
        except Exception:
            continue
        for r in rows:
            for field in ("物流单号", "快递信息", "快递单号"):
                text = norm(p(r, field))
                for match in re.findall(r"SF[A-Z0-9]+", text, flags=re.I):
                    out.add(match.upper())
    return out


def extract_skus(raw: dict, year_month: str) -> set[str]:
    out = {norm(s).upper() for s in raw.get("sku_set", set()) if norm(s)}
    for sf in raw.get("source_files", []):
        fname = _fname(sf)
        if not fname.lower().endswith((".xlsx", ".xls", ".csv")):
            continue
        if not _is_order_detail_export(fname):
            continue
        try:
            rows = sheet_rows(sf.get("buf") or b"", fname, "export") or sheet_rows(sf.get("buf") or b"", fname)
        except Exception:
            continue
        for r in rows:
            sku = norm(p(r, "商家编码") or p(r, "外部系统编号") or p(r, "商家编码-规格维度"))
            if sku:
                out.add(sku.upper())
    return out


def read_tmall_fees(report: SettlementReport, files: list[dict[str, Any]], shop: str,
                    fee_by_order: dict[str, float]) -> tuple[float, float]:
    platform_fee = 0.0
    ad_fee = 0.0
    ym = report.year_month
    for sf in sorted(files, key=lambda x: _fname(x)):
        fname = _fname(sf)
        base = _basename(fname)
        if not fname.lower().endswith(".csv") or base.startswith("~$"):
            continue
        if "交易货款" in fname:
            continue
        rows = read_csv(sf.get("buf") or b"", prefer_gbk=True)
        if base.startswith("0947_"):
            for r in rows:
                trade_date = norm(p(r, "交易日期"))
                typ = norm(p(r, "交易类型"))
                direction = norm(p(r, "收支类型"))
                amount = money(p(r, "操作金额(元)"))
                include = trade_date.startswith(ym) and (typ == "扣款" or direction == "支出")
                report.add_fee("天猫", shop, fname, "广告费", "", amount, "操作金额(元)",
                               "计入：交易日期为当月且扣款/支出" if include else "排除：充值或非当月交易日期")
                if include:
                    ad_fee += amount
            continue
        if "流水" in base:
            continue
        for r in rows:
            amount = 0.0
            field = ""
            for candidate in ("扣费金额(元)", "扣费金额", "本月付款", "积分类服务费金额", "账单金额"):
                amount = money(p(r, candidate))
                if amount:
                    field = candidate
                    break
            if not amount:
                continue
            order_id = ""
            for candidate in ("订单号", "交易主订单号", "交易主单号"):
                order_id = norm(p(r, candidate))
                if order_id:
                    break
            cat = re.sub(r"_202606_202606\.csv$|\.csv$", "", base)
            platform_fee += amount
            report.add_fee("天猫", shop, fname, cat, order_id, amount, field,
                           "计入毛利平台/其他费用；排除-流水文件避免重复")
            if order_id:
                fee_by_order[order_id] += amount
    return platform_fee, ad_fee


def process_tmall(report: SettlementReport, raw: dict, cost_map: dict[str, dict[str, Any]],
                  bill_pool: dict[str, dict[str, Any]], shop: str) -> None:
    files = _files(raw, "天猫", shop)
    if not files:
        return
    # Prefer shop-scoped logistics where present. Some operators upload the
    # carrier month bills only once in the global company slot, so miss in the
    # shop pool must fall back to the global pool passed by compute().
    shop_bill_pool = build_bill_pool({"source_files": files, "logistics": []})
    trade_sf = _pick_file(files, "交易货款")
    if not trade_sf:
        report.add_gap("P0", "天猫", shop, "资料缺口", shop, "未找到交易货款结算文件", "无法计算天猫结算月收入",
                       "补交易货款文件后重跑")
        report.add_monthly("天猫", shop, 0, 0, 0, 0, 0, 0, 0, 0, 0, "缺交易货款文件")
        return
    order_files = [sf for sf in files if _is_order_detail_export(_fname(sf))]
    trade_rows = read_csv(trade_sf.get("buf") or b"", prefer_gbk=True)
    order_rows: list[dict[str, Any]] = []
    for sf in order_files:
        rows = sheet_rows(sf.get("buf") or b"", _fname(sf), "export") or sheet_rows(sf.get("buf") or b"", _fname(sf))
        order_rows.extend(rows)
        report.add_source("天猫", shop, "订单明细", _fname(sf), len(rows), "已读取",
                          "补ERP_SKU/物流单号；支持上月+本月多文件合并")
    report.add_source("天猫", shop, "交易货款", _fname(trade_sf), len(trade_rows), "已读取", "结算月收入/退款主来源")

    order_by_sub: dict[str, dict[str, Any]] = {}
    for r in order_rows:
        sub = norm(p(r, "子订单编号"))
        if sub:
            merged = dict(order_by_sub.get(sub, {}))
            # 淘宝可能把物流字段放在 ExportOrderList、商家编码放在
            # ExportItemlList；按子订单合并非空字段，避免只读到半张信息。
            merged.update({key: value for key, value in r.items() if norm(value)})
            order_by_sub[sub] = merged
    fee_by_order: dict[str, float] = defaultdict(float)
    platform_fee, ad_fee = read_tmall_fees(report, files, shop, fee_by_order)
    report.add_source("天猫", shop, "广告/平台费用", "多个csv", "", "已读取",
                      "广告取0947扣款；平台费用排除-流水文件避免重复")
    report.add_source("天猫", shop, "物流账单池", "顺丰/中通 前后月账单+API", len(shop_bill_pool), "已读取",
                      "按运单号匹配，顺丰缺口尝试API")

    lines: list[dict[str, Any]] = []
    waybill_groups: dict[str, list[int]] = defaultdict(list)
    for r in trade_rows:
        sub = norm(p(r, "子订单号"))
        order = order_by_sub.get(sub)
        main = norm(p(r, "订单号"))
        sku = norm(p(order, "商家编码") or p(order, "外部系统编号"))
        name = norm(p(order, "商品标题") or p(r, "商品名称"))
        qty = money(p(r, "数量"))
        sales = money(p(r, "订单实际金额（元）"))
        refund = money(p(r, "退款金额（元）"))
        refund_qty = min(qty, qty * refund / sales) if sales > 0 and refund > 0 else 0.0
        net_qty = max(qty - refund_qty, 0.0)
        unit, cost_src, cost_name = cost_entry(cost_map, sku)
        if not name:
            name = cost_name
        purchase = net_qty * unit
        if not sku:
            report.add_gap("P0", "天猫", shop, "采购成本", main, "订单无法取得商家编码/外部系统编号",
                           "无法映射采购成本", "补订单明细商家编码或SKU对照表")
        elif net_qty > 0 and unit <= 0:
            report.add_gap("P0", "天猫", shop, "采购成本", sku, "采购成本表未匹配或成本为0",
                           "毛利会虚高", "维护产品采购成本台后重跑")
        report.add_cost("天猫", shop, main, sku, name, net_qty, unit, purchase, cost_src)
        waybill = norm(p(order, "物流单号")).removeprefix("No:")
        carrier = norm(p(order, "物流公司"))
        line = {
            "order": main, "sub": sub, "sku": sku, "name": name, "qty": qty,
            "refund_qty": refund_qty, "sales": sales, "refund": refund,
            "purchase": purchase, "waybill": waybill, "carrier": carrier, "tail": 0.0,
        }
        lines.append(line)
        if waybill:
            waybill_groups[waybill].append(len(lines) - 1)
    for waybill, idxs in waybill_groups.items():
        bill = shop_bill_pool.get(waybill) or bill_pool.get(waybill)
        if bill:
            share = float(bill["amount"]) / len(idxs) if idxs else 0.0
            for idx in idxs:
                lines[idx]["tail"] = share
                report.add_log("天猫", shop, lines[idx]["order"], lines[idx]["sub"], waybill,
                               lines[idx]["carrier"], share, bill["file"], bill["sheet"],
                               bill["source"], "计入-结算订单运单命中")
        else:
            for idx in idxs:
                report.add_log("天猫", shop, lines[idx]["order"], lines[idx]["sub"], waybill,
                               lines[idx]["carrier"], 0, "", "", "未命中", "P0-物流缺口")
                report.add_gap("P0", "天猫", shop, "物流成本", waybill,
                               "结算订单运单未在前后月账单池命中，顺丰API也未返回费用",
                               "尾程费用缺失", "补后续账单或核实运单/API权限后重跑")
    for line in lines:
        if not line["waybill"]:
            report.add_log("天猫", shop, line["order"], line["sub"], "", line["carrier"],
                           0, "", "", "缺物流单号", "P0-物流缺口")
            report.add_gap("P0", "天猫", shop, "物流成本", line["order"],
                           "结算订单订单明细物流单号为空", "无法匹配尾程费用", "补完整订单明细/物流单号")

    sales_total = sum(float(x["sales"]) for x in lines)
    refund_total = sum(float(x["refund"]) for x in lines)
    qty_total = sum(float(x["qty"]) for x in lines)
    purchase_total = sum(float(x["purchase"]) for x in lines)
    tail_total = sum(float(x["tail"]) for x in lines)
    net_total = sales_total - refund_total
    for line in lines:
        net = float(line["sales"]) - float(line["refund"])
        ratio = net / net_total if net_total else 0.0
        report.add_product("天猫", shop, line["sku"], line["name"], float(line["qty"]),
                           float(line["refund_qty"]), float(line["sales"]), float(line["refund"]),
                           platform_fee * ratio, ad_fee * ratio, float(line["purchase"]),
                           float(line["tail"]), 0.0)
    status = "可作为自动化对比基准" if report.gap_count("天猫", shop, "P0") == 0 else "存在P0缺口-待补后重跑"
    report.add_monthly("天猫", shop, len(lines), qty_total, sales_total, refund_total, platform_fee,
                       ad_fee, purchase_total, tail_total, 0.0, status)


def add_dy_index(target: dict[str, list[dict[str, Any]]], key: str, row: dict[str, Any]) -> None:
    if key:
        target.setdefault(key, []).append(row)


def pick_dy_lines(by_order_product: dict[str, list[dict[str, Any]]],
                  by_order: dict[str, list[dict[str, Any]]],
                  order_id: str, product_id: str) -> list[dict[str, Any]]:
    return by_order_product.get(f"{order_id}|{product_id}") or by_order.get(order_id) or []


def parse_douyin_express(info: Any) -> list[dict[str, Any]]:
    text = norm(info)
    if not text or text == "-":
        return []
    out: list[dict[str, Any]] = []
    for part in text.split(";"):
        part = norm(part)
        if not part:
            continue
        cols = [norm(c) for c in part.split(",")]
        main = cols[0]
        waybill = main
        carrier = ""
        if "-" in main:
            waybill, carrier = [norm(x) for x in main.split("-", 1)]
        product_id = ""
        qty = 1.0
        if len(cols) >= 3:
            product_id = norm(cols[-2].split("-")[-1])
            qty = money(cols[-1]) or 1.0
        if not carrier:
            if waybill.upper().startswith("SF"):
                carrier = "顺丰速运"
            elif waybill.upper().startswith("JT"):
                carrier = "极兔速递"
            elif re.fullmatch(r"\d+", waybill):
                carrier = "中通快递"
        if waybill:
            out.append({"waybill": waybill, "carrier": carrier, "product_id": product_id, "qty": qty})
    return out


def process_douyin(report: SettlementReport, raw: dict, cost_map: dict[str, dict[str, Any]],
                   bill_pool: dict[str, dict[str, Any]], shop: str) -> None:
    files = _files(raw, "抖音", shop)
    if not files:
        return
    settle_sf = _pick_file(files, "结算订单")
    order_sf = _pick_file(files, "订单明细")
    if not settle_sf or not order_sf:
        report.add_gap("P0", "抖音", shop, "资料缺口", shop, "缺结算订单或订单明细文件",
                       "无法计算销售/成本/物流", "补结算订单和订单明细后重跑")
        report.add_monthly("抖音", shop, 0, 0, 0, 0, 0, 0, 0, 0, 0, "缺结算订单或订单明细")
        return
    settle_rows = [r for r in read_csv(settle_sf.get("buf") or b"") if norm(p(r, "订单号"))]
    order_rows = sheet_rows(order_sf.get("buf") or b"", _fname(order_sf), "Sheet1")
    report.add_source("抖音", shop, "结算订单", _fname(settle_sf), len(settle_rows), "已读取", "按结算时间/结算月")
    report.add_source("抖音", shop, "订单明细", _fname(order_sf), len(order_rows), "已读取", "补ERP_SKU/物流单号/快递信息")
    by_order_product: dict[str, list[dict[str, Any]]] = {}
    by_order: dict[str, list[dict[str, Any]]] = {}
    for r in order_rows:
        oid = norm(p(r, "主订单编号"))
        product_id = norm(p(r, "商品ID"))
        add_dy_index(by_order, oid, r)
        add_dy_index(by_order_product, f"{oid}|{product_id}", r)

    other_fee = 0.0
    for sf in _matching_files(files, "动账"):
        tx_rows = read_csv(sf.get("buf") or b"")
        report.add_source("抖音", shop, "动账", _fname(sf), len(tx_rows), "已读取",
                          "只纳入权益保险/消费者赔付/上门取件运费等经营费用")
        for r in tx_rows:
            scene = norm(p(r, "动账场景"))
            direction = norm(p(r, "动账方向"))
            amount = money(p(r, "动账金额"))
            if scene in ("权益保险", "消费者赔付", "上门取件运费") and direction == "出账":
                fee = abs(amount)
                other_fee += fee
                report.add_fee("抖音", shop, _fname(sf), scene, norm(p(r, "订单号")), fee, "动账金额",
                               "计入其他经营费用")
            elif norm(p(r, "支付保费")):
                fee = money(p(r, "支付保费"))
                if fee > 0:
                    other_fee += fee
                    report.add_fee("抖音", shop, _fname(sf), "权益保险", norm(p(r, "订单编号")), fee, "支付保费",
                                   "计入其他经营费用")
    for sf in files:
        fname = _fname(sf)
        if "税务" in fname or "涉税" in fname:
            rows = read_csv(sf.get("buf") or b"")
            count = sum(1 for r in rows if norm(p(r, "报送场景")) or norm(p(r, "订单号")))
            report.add_source("抖音", shop, "涉税信息", fname, count, "已读取",
                              "用于后续税务A/B核对；本月毛利不重复计销售服务收入")
            report.add_tax("抖音", shop, fname, "涉税信息", count, "已读取", "已读入涉税资料，P0未重复计入毛利收入")
        if fname.endswith(".txt") and "广告" in fname:
            try:
                note = (sf.get("buf") or b"").decode("utf-8", errors="ignore")
            except Exception:
                note = ""
            report.add_source("抖音", shop, "广告说明", fname, "", "已读取", norm(note))

    other_base = sum(money(p(r, "收入合计")) for r in settle_rows if norm(p(r, "结算单类型")) == "已结算")
    sales_total = refund_total = qty_total = platform_fee = purchase_total = tail_total = ad_fee = payback = 0.0
    counted_waybills: set[str] = set()
    for r in settle_rows:
        order_id = norm(p(r, "订单号"))
        product_id = norm(p(r, "商品ID"))
        typ = norm(p(r, "结算单类型"))
        qty = money(p(r, "商品数量"))
        income = money(p(r, "收入合计"))
        settle = money(p(r, "结算金额"))
        pf = abs(money(p(r, "平台服务费")))
        offsite = abs(money(p(r, "站外推广费")))
        platform_fee += pf + offsite
        payback += settle
        if pf:
            report.add_fee("抖音", shop, _fname(settle_sf), "平台服务费", order_id, pf, "平台服务费", "计入平台费用")
        if offsite:
            report.add_fee("抖音", shop, _fname(settle_sf), "站外推广费", order_id, offsite, "站外推广费", "计入平台费用")
        is_income = typ == "已结算"
        is_refund = ("退款" in typ) or income < 0 or settle < 0
        if is_income:
            qty_total += qty
            sales_total += income
        elif is_refund:
            refund_total += abs(income if income else settle)
        lines = pick_dy_lines(by_order_product, by_order, order_id, product_id)
        skus = sorted({norm(p(x, "商家编码")) for x in lines if norm(p(x, "商家编码"))})
        sku = skus[0] if skus else ""
        name = norm(p(lines[0], "选购商品")) if lines else norm(p(r, "商品名称"))
        if not lines:
            report.add_gap("P0", "抖音", shop, "订单明细匹配", order_id, "结算订单未在订单明细中找到",
                           "无法补商家编码和物流单号", "补覆盖该订单下单时间范围的订单明细后重跑")
        elif len(skus) > 1:
            report.add_gap("P0", "抖音", shop, "采购成本", order_id,
                           f"同一结算订单匹配多个商家编码：{','.join(skus)}", "无法唯一映射采购成本",
                           "运营核对订单明细商品ID/商家编码关系")
        elif not sku and (is_income or is_refund):
            report.add_gap("P0", "抖音", shop, "采购成本", order_id, "订单明细商家编码为空",
                           "无法映射采购成本", "补商家编码/ERP_SKU")
        net_qty = qty if is_income else (-abs(qty) if is_refund else 0.0)
        unit, cost_src, cost_name = cost_entry(cost_map, sku)
        if not name:
            name = cost_name
        purchase = max(net_qty * unit, 0.0)
        if (is_income or is_refund) and net_qty > 0:
            if sku and unit <= 0:
                report.add_gap("P0", "抖音", shop, "采购成本", sku, "采购成本表未匹配或成本为0",
                               "毛利会虚高", "维护产品采购成本台后重跑")
            purchase_total += purchase
            report.add_cost("抖音", shop, order_id, sku, name, net_qty, unit, purchase, cost_src)
        tail_for_order = 0.0
        if is_income:
            express_items: list[dict[str, Any]] = []
            for line in lines:
                express_items.extend(parse_douyin_express(p(line, "快递信息")))
            if not express_items:
                report.add_gap("P0", "抖音", shop, "物流成本", order_id, "结算收入订单未解析到有效运单号",
                               "尾程费用缺失", "补订单明细快递信息/物流单号")
                report.add_log("抖音", shop, order_id, "", "", "", 0, "", "", "未解析到运单", "P0-物流缺口")
            local: set[str] = set()
            for ex in express_items:
                wb = ex["waybill"]
                if not wb or wb in local:
                    continue
                local.add(wb)
                bill = bill_pool.get(wb)
                if bill:
                    counted = wb not in counted_waybills
                    amount = float(bill["amount"]) if counted else 0.0
                    if counted:
                        counted_waybills.add(wb)
                        tail_total += amount
                        tail_for_order += amount
                    report.add_log("抖音", shop, order_id, "", wb, ex["carrier"], amount,
                                   bill["file"], bill["sheet"], bill["source"],
                                   "计入-运单命中" if counted else "重复运单不重复计费")
                else:
                    report.add_log("抖音", shop, order_id, "", wb, ex["carrier"], 0, "", "", "未命中", "P0-物流缺口")
                    report.add_gap("P0", "抖音", shop, "物流成本", wb,
                                   f"{ex['carrier']} 运单未在前后月账单池命中，顺丰API也未返回费用",
                                   "尾程费用缺失", "补后续账单或核实运单/API权限后重跑")
        if is_income or is_refund:
            net = income if is_income else -abs(income if income else settle)
            line_other = other_fee * income / other_base if is_income and other_base else 0.0
            report.add_product("抖音", shop, sku, name, qty if is_income else 0.0,
                               abs(qty) if is_refund else 0.0, income if is_income else 0.0,
                               abs(net) if is_refund else 0.0, pf + offsite, 0.0, purchase,
                               tail_for_order, line_other)
    status = "可作为自动化对比基准" if report.gap_count("抖音", shop, "P0") == 0 else "存在P0缺口-待补后重跑"
    report.add_monthly("抖音", shop, len(settle_rows), qty_total, sales_total, refund_total,
                       platform_fee, ad_fee, purchase_total, tail_total, other_fee, status)


def process_xhs(report: SettlementReport, raw: dict, cost_map: dict[str, dict[str, Any]],
                bill_pool: dict[str, dict[str, Any]], shop: str) -> None:
    files = _files(raw, "小红书", shop)
    if not files:
        return
    settle_sf = _pick_file(files, "商品结算明细")
    order_sf = _pick_file(files, "小红书订单查询") or _pick_file(files, "订单查询") or _pick_file(files, "结算订单明细")
    settle_rows: list[dict[str, Any]] = []
    if settle_sf:
        settle_rows = sheet_rows(settle_sf.get("buf") or b"", _fname(settle_sf), "商品结算明细")
        report.add_source("小红书", shop, "商品结算明细", _fname(settle_sf), len(settle_rows), "已读取",
                          "结算月收入/退款主来源")
    else:
        no_settle = [sf for sf in files if _fname(sf).endswith(".txt") and "无结算" in _fname(sf)]
        if no_settle:
            report.add_source("小红书", shop, "结算说明", _fname(no_settle[0]), "", "已读取",
                              norm((no_settle[0].get("buf") or b"").decode("utf-8", errors="ignore")))
        else:
            report.add_gap("P0", "小红书", shop, "资料缺口", shop, "缺商品结算明细或无结算订单说明",
                           "无法确认本月结算收入/退款", "补商品结算明细或无结算订单说明")
    order_rows: list[dict[str, Any]] = []
    if order_sf:
        order_rows = sheet_rows(order_sf.get("buf") or b"", _fname(order_sf), "包裹详情")
        report.add_source("小红书", shop, "订单查询", _fname(order_sf), len(order_rows), "已读取",
                          "补ERP_SKU/物流单号")
    elif settle_rows:
        report.add_gap("P0", "小红书", shop, "资料缺口", shop, "缺订单查询/包裹详情文件",
                       "无法补商家编码和物流单号", "补小红书订单查询导出")
    for sf in files:
        fname = _fname(sf)
        if "税务" in fname:
            rows = sheet_rows(sf.get("buf") or b"", fname)
            report.add_source("小红书", shop, "涉税信息", fname, len(rows) if rows else "", "已读取",
                              "用于后续税务A/B核对；不重复计入毛利收入")
            report.add_tax("小红书", shop, fname, "涉税信息", len(rows) if rows else "", "已读取",
                           "已读入涉税资料，P0未重复计入毛利收入")
        if fname.endswith(".txt") and "广告" in fname:
            report.add_source("小红书", shop, "广告说明", fname, "", "已读取",
                              norm((sf.get("buf") or b"").decode("utf-8", errors="ignore")))
    order_by_order = {norm(p(r, "订单号")): r for r in order_rows if norm(p(r, "订单号"))}
    other_fee = 0.0
    for sf in files:
        fname = _fname(sf)
        if "支出" not in fname and "平台支出" not in fname:
            continue
        rows = sheet_rows(sf.get("buf") or b"", fname, "资金明细")
        report.add_source("小红书", shop, "平台支出", fname, len(rows), "已读取", "动账方向=支出计入其他费用")
        for r in rows:
            direction = norm(p(r, "动账方向"))
            scene = norm(p(r, "动账场景"))
            amt = money(p(r, "动账金额"))
            if direction == "支出" or amt < 0:
                fee = abs(amt)
                other_fee += fee
                report.add_fee("小红书", shop, fname, scene, norm(p(r, "业务单号")), fee, "动账金额", "计入其他费用")
    sales_total = refund_total = qty_total = platform_fee_signed = purchase_total = tail_total = ad_fee = 0.0
    counted_waybills: set[str] = set()
    product_lines: list[dict[str, Any]] = []
    for r in settle_rows:
        order_id = norm(p(r, "订单号"))
        if not order_id:
            continue
        txn_type = norm(p(r, "交易类型"))
        is_income = txn_type == "结算入账"
        is_refund = txn_type == "退款"
        qty = money(p(r, "商品数量"))
        cash = money(p(r, "商品实付/实退")) or money(p(r, "计佣基数"))
        platform_fee_signed += money(p(r, "佣金总额")) + money(p(r, "分销佣金")) + money(p(r, "花呗分期手续费"))
        order = order_by_order.get(order_id)
        sku = norm(p(order, "商家编码"))
        name = norm(p(order, "商品名称") or p(r, "商品名称"))
        order_qty = money(p(order, "SKU件数"))
        if (is_income or is_refund) and qty == 0 and order_qty:
            qty = order_qty
        if is_income:
            sales_total += abs(cash)
            qty_total += qty
        elif is_refund:
            refund_total += abs(cash)
        if order is None:
            report.add_gap("P0", "小红书", shop, "订单明细匹配", order_id, "商品结算订单未在订单查询中找到",
                           "无法补商家编码和物流单号", "补覆盖该订单下单时间范围的订单查询")
        elif not sku and (is_income or is_refund):
            report.add_gap("P0", "小红书", shop, "采购成本", order_id, "订单查询商家编码为空",
                           "无法映射采购成本", "补商家编码/ERP_SKU")
        unit, cost_src, cost_name = cost_entry(cost_map, sku)
        if not name:
            name = cost_name
        if is_income and sku and unit <= 0:
            report.add_gap("P0", "小红书", shop, "采购成本", sku, "采购成本表未匹配或成本为0",
                           "毛利会虚高", "维护产品采购成本台后重跑")
        signed_qty = qty if is_income else (-abs(qty) if is_refund else 0.0)
        purchase = signed_qty * unit
        purchase_total += purchase
        if is_income or is_refund:
            report.add_cost("小红书", shop, order_id, sku, name, signed_qty, unit, purchase, cost_src)
        tail_for_order = 0.0
        if is_income and order is not None:
            wb = norm(p(order, "快递单号"))
            carrier = norm(p(order, "快递公司"))
            if not wb:
                report.add_gap("P0", "小红书", shop, "物流成本", order_id, "结算收入订单快递单号为空",
                               "尾程费用缺失", "补订单查询快递单号")
                report.add_log("小红书", shop, order_id, "", "", carrier, 0, "", "", "缺快递单号", "P0-物流缺口")
            else:
                bill = bill_pool.get(wb)
                if bill:
                    counted = wb not in counted_waybills
                    amount = float(bill["amount"]) if counted else 0.0
                    if counted:
                        counted_waybills.add(wb)
                        tail_total += amount
                        tail_for_order += amount
                    report.add_log("小红书", shop, order_id, "", wb, carrier, amount,
                                   bill["file"], bill["sheet"], bill["source"],
                                   "计入-运单命中" if counted else "重复运单不重复计费")
                else:
                    report.add_log("小红书", shop, order_id, "", wb, carrier, 0, "", "", "未命中", "P0-物流缺口")
                    report.add_gap("P0", "小红书", shop, "物流成本", wb,
                                   f"{carrier} 运单未在前后月账单池命中，顺丰API也未返回费用",
                                   "尾程费用缺失", "补后续账单或核实运单/API权限后重跑")
        if is_income or is_refund:
            product_lines.append({
                "sku": sku, "name": name, "is_income": is_income, "is_refund": is_refund,
                "qty": qty, "sales": abs(cash) if is_income else 0.0,
                "refund": abs(cash) if is_refund else 0.0, "purchase": purchase,
                "tail": tail_for_order,
            })
    platform_fee = abs(platform_fee_signed)
    net_sales = sales_total - refund_total
    for pl in product_lines:
        line_net = pl["sales"] - pl["refund"]
        ratio = line_net / net_sales if net_sales else 0.0
        report.add_product("小红书", shop, pl["sku"], pl["name"], pl["qty"] if pl["is_income"] else 0.0,
                           pl["qty"] if pl["is_refund"] else 0.0, pl["sales"], pl["refund"],
                           platform_fee * ratio, 0.0, pl["purchase"], pl["tail"], other_fee * ratio)
    status = "可作为自动化对比基准" if report.gap_count("小红书", shop, "P0") == 0 else "存在P0缺口-待补后重跑"
    report.add_monthly("小红书", shop, len(settle_rows), qty_total, sales_total, refund_total,
                       platform_fee, ad_fee, purchase_total, tail_total, other_fee, status)


def process_jd(report: SettlementReport, raw: dict, shop: str) -> None:
    files = _files(raw, "京东", shop)
    if not files:
        report.add_source("京东", shop, "资料文件", "", "", "缺失",
                          "未读取到订单结算明细/到账文件，也没有结构化无结算确认")
        report.add_gap("P0", "京东", shop, "资料缺口", shop,
                       "缺订单结算明细/到账文件或无结算确认", "无法确认本月京东结算收入/费用",
                       "补订单结算明细、到账文件，或在卡片确认本月无结算")
        report.add_monthly("京东", shop, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                           "缺资料/无结算确认-待运营补证据")
        return
    settle_files = [
        sf for sf in files
        if "订单结算明细对账" in _fname(sf) or "货款明细" in _fname(sf) or "结算明细" in _fname(sf)
    ]
    platform_fee = 0.0
    read_any = False
    for sf in settle_files:
        fname = _fname(sf)
        rows = read_csv(sf.get("buf") or b"")
        read_any = True
        report.add_source("京东", shop, "订单结算明细对账", fname, len(rows), "已读取",
                          "按到账时间/账单生成时间筛当月")
        for r in rows:
            arrived = norm(p(r, "到账时间") or p(r, "账单生成时间") or p(r, "结算时间"))
            if report.year_month.replace("-", "") not in arrived.replace("-", "").replace("/", "") and not arrived.startswith(report.year_month):
                continue
            direction = norm(p(r, "收支方向"))
            amount = money(p(r, "应结金额") or p(r, "金额") or p(r, "结算金额"))
            fee_name = norm(p(r, "费用名称") or p(r, "业务类型") or p(r, "费用类型"))
            if direction == "支出" or amount < 0:
                if any(x in fee_name for x in ("货款", "提现", "保证金", "充值", "转账")):
                    continue
                fee = abs(amount)
                platform_fee += fee
                report.add_fee("京东", shop, fname, fee_name or "京东结算支出",
                               norm(p(r, "订单编号") or p(r, "订单号")), fee,
                               "应结金额/收支方向=支出", "京东结算月支出，当前无订单明细，计入平台/其他费用")
    for sf in files:
        fname = _fname(sf)
        if fname.endswith(".txt"):
            report.add_source("京东", shop, "资料文件", fname, "", "已登记",
                              norm((sf.get("buf") or b"").decode("utf-8", errors="ignore")))
    if read_any or platform_fee:
        status = "无订单数据，仅有结算支出，作为自动化对比基准" if platform_fee else "运营提供无订单/无结算确认，按0试算"
        report.add_monthly("京东", shop, 0, 0, 0, 0, platform_fee, 0, 0, 0, 0, status)
    else:
        report.add_monthly("京东", shop, 0, 0, 0, 0, 0, 0, 0, 0, 0, "运营提供无订单/无结算确认，按0试算")


def compute(raw: dict, cost_map: dict[str, dict[str, Any]], year_month: str) -> dict[str, Any]:
    report = SettlementReport(year_month)
    bill_pool = build_bill_pool(raw)
    for shop in ("天猫纷岚", "天猫宝空"):
        process_tmall(report, raw, cost_map, bill_pool, shop)
    for shop in ("抖音宝空", "抖音纷岚"):
        process_douyin(report, raw, cost_map, bill_pool, shop)
    for shop in ("小红书宝空", "小红书纷岚"):
        process_xhs(report, raw, cost_map, bill_pool, shop)
    for shop in ("京东纷岚", "京东宝空"):
        process_jd(report, raw, shop)
    return report.as_dict()
