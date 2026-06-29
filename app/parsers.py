"""解析平台报表附件 — 第一版只支持天猫 POWKONG 旗舰店格式."""
import io
import csv
import re
import datetime
from collections import Counter, defaultdict
import openpyxl
import xlrd


def _is_apr(d, year_month: str) -> bool:
    """判断日期是否在指定 YYYY-MM."""
    if d is None:
        return False
    if isinstance(d, datetime.datetime):
        return d.strftime("%Y-%m") == year_month
    return year_month in str(d)


def _load_rows(buf: bytes, filename: str = "") -> list[tuple]:
    """统一读取 .xlsx/.xls/.csv → list[tuple](含表头行). 按文件 magic 探测, 不靠扩展名."""
    if not buf:
        return []
    head = buf[:8]
    if head[:4] == b"PK\x03\x04":  # xlsx (zip)
        # 非 read_only: 部分淘宝/天猫 xlsx 的 dimension 不准, read_only 模式会只读到 1 行
        wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
        rows = [tuple(r) for r in wb.active.iter_rows(values_only=True)]
        wb.close()
        return rows
    if head[:4] == b"\xd0\xcf\x11\xe0":  # 老 .xls (OLE2)
        wb = xlrd.open_workbook(file_contents=buf)
        s = wb.sheet_by_index(0)
        out = []
        for i in range(s.nrows):
            row = []
            for j in range(s.ncols):
                c = s.cell(i, j)
                if c.ctype == xlrd.XL_CELL_DATE:  # Excel 日期序列号 → datetime
                    try:
                        row.append(xlrd.xldate.xldate_as_datetime(c.value, wb.datemode))
                    except Exception:
                        row.append(c.value)
                else:
                    row.append(c.value)
            out.append(tuple(row))
        return out
    for enc in ("gbk", "utf-8-sig", "utf-8"):  # 其余按 CSV 文本(淘宝导出多为 gbk)
        try:
            return [tuple(r) for r in csv.reader(io.StringIO(buf.decode(enc)))]
        except Exception:
            continue
    return []


def _pick(col: dict, *names):
    """按列名别名返回第一个命中的列 idx, 都没有返回 -1."""
    for n in names:
        if n in col:
            return col[n]
    return -1


def _ym_match(t, year_month: str) -> bool:
    """鲁棒年月匹配: 兼容 datetime / '2026-01-15' / '2026/1/29 12:04' / '2026年1月' 等."""
    if t is None or t == "":
        return False
    if isinstance(t, datetime.datetime):
        return t.strftime("%Y-%m") == year_month
    m = re.search(r"(\d{4})[/\-年.](\d{1,2})", str(t))
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}" == year_month
    return year_month in str(t)


# ===== 天猫订单明细 =====
def parse_tmall_orders(buf: bytes, year_month: str, filename: str = "") -> tuple[list[dict], set]:
    """解析天猫/淘宝订单明细 (.xlsx/.xls/.csv) → (rows, sku_set). 兼容新老两套列名."""
    all_rows = _load_rows(buf, filename)
    if not all_rows:
        return [], set()
    header = all_rows[0]
    col = {h: i for i, h in enumerate(header) if h not in (None, "")}
    i_pay = _pick(col, "订单付款时间")
    i_create = _pick(col, "订单创建时间")
    i_sku = _pick(col, "商家编码")
    i_main = _pick(col, "主订单编号")
    i_sub = _pick(col, "子订单编号")
    i_title = _pick(col, "商品标题", "标题")
    i_attr = _pick(col, "商品属性")
    i_qty = _pick(col, "购买数量")
    i_price = _pick(col, "商品价格", "价格")
    i_paid = _pick(col, "买家实付金额", "买家实际支付金额")
    i_track = _pick(col, "物流单号")
    i_status = _pick(col, "订单状态")
    i_rstatus = _pick(col, "退款状态")
    i_payable = _pick(col, "买家应付货款")
    i_order_refund = _pick(col, "退款金额")  # 订单表退款金额(退款总额含券), 仅天猫/淘宝有
    out = []
    sku_set = set()

    def g(r, i):
        return r[i] if 0 <= i < len(r) else None

    for r in all_rows[1:]:
        pay = g(r, i_pay)
        create = g(r, i_create)
        t = pay or create
        if not _ym_match(t, year_month):
            continue
        sku = g(r, i_sku) or ""
        if sku:
            sku_set.add(sku)
        out.append({
            "main_oid": g(r, i_main) or "",
            "sub_oid": g(r, i_sub) or "",
            "create_t": str(create or ""),
            "pay_t": str(t),
            "sku": sku,
            "title": g(r, i_title) or "",
            "attr": g(r, i_attr) or "",
            "qty": g(r, i_qty) or 0,
            "price": g(r, i_price) or 0,
            "paid": g(r, i_paid) or 0,
            "tracking": g(r, i_track) or "",
            "status": g(r, i_status) or "",
            "refund_status": g(r, i_rstatus) or "",
            "payable": g(r, i_payable) or 0,
            "order_refund": g(r, i_order_refund) or 0,
        })
    return out, sku_set


# ===== 天猫退款明细 =====
def parse_tmall_refunds(buf: bytes, filename: str = "") -> list[dict]:
    """天猫/淘宝退款. A=真退款表(有'退款总额'+'退款编号'); B=运营误传订单表(从'退款金额'>0 行提取, 淘宝C店1-3月)."""
    all_rows = _load_rows(buf, filename)
    if not all_rows:
        return []
    header = all_rows[0]
    col = {h: i for i, h in enumerate(header) if h not in (None, "")}

    def g(r, i):
        return r[i] if 0 <= i < len(r) else None

    out = []
    i_total = _pick(col, "退款总额", "退款金额（元）")
    # 情形 A: 真退款明细表
    if i_total >= 0 and _pick(col, "退款编号") >= 0:
        i_rid = _pick(col, "退款编号"); i_main = _pick(col, "订单编号")
        i_sku = _pick(col, "商家编码"); i_title = _pick(col, "宝贝标题", "商品标题")
        i_ct = _pick(col, "退款完结时间"); i_type = _pick(col, "售后类型")
        i_reason = _pick(col, "买家退款原因"); i_payt = _pick(col, "订单付款时间")
        i_tb = _pick(col, "退给买家金额"); i_tp = _pick(col, "退给平台金额")
        i_status = _pick(col, "退款状态")  # 退款成功/退款关闭(换货等不真退) — 引擎只认成功
        for r in all_rows[1:]:
            out.append({
                "refund_id": g(r, i_rid) or "", "main_oid": g(r, i_main) or "",
                "sku": g(r, i_sku) or "", "title": g(r, i_title) or "",
                "complete_t": str(g(r, i_ct) or ""), "amount": g(r, i_total) or 0,
                "type": g(r, i_type) or "", "reason": g(r, i_reason) or "",
                "to_buyer": g(r, i_tb) or 0, "to_platform": g(r, i_tp) or 0,
                "pay_t": str(g(r, i_payt) or ""), "status": g(r, i_status) or "",
            })
        return out
    # 情形 B: 运营传的是订单明细表 → 从"退款金额">0 的行提取
    i_amt = _pick(col, "退款金额")
    if i_amt < 0:
        return []
    i_sku = _pick(col, "商家编码"); i_main = _pick(col, "主订单编号", "订单编号")
    i_title = _pick(col, "商品标题", "标题"); i_rstatus = _pick(col, "退款状态")
    i_payt = _pick(col, "订单付款时间")
    for r in all_rows[1:]:
        try:
            amt = float(g(r, i_amt) or 0)
        except (TypeError, ValueError):
            amt = 0
        if amt <= 0:
            continue
        out.append({
            "refund_id": "", "main_oid": g(r, i_main) or "", "sku": g(r, i_sku) or "",
            "title": g(r, i_title) or "", "complete_t": "", "amount": amt,
            "type": g(r, i_rstatus) or "", "reason": "", "to_buyer": amt, "to_platform": 0,
            "pay_t": str(g(r, i_payt) or ""),
        })
    return out


# ===== 天猫平台费用 (单文件，月度汇总粒度) =====
def parse_tmall_platform_fee(buf: bytes, source_name: str, filename: str = "") -> list[dict]:
    """天猫/淘宝平台费. 支持 .xlsx/.xls/.csv. 淘宝分项 csv 每行一笔, 真实费列='扣费金额'(明细累加)."""
    all_rows = _load_rows(buf, filename)
    if len(all_rows) < 2:
        return []
    header = all_rows[0]
    col = {str(h).strip(): i for i, h in enumerate(header) if h not in (None, "")}
    # "扣费金额" 在最前: 淘宝分项 csv 明细费列(非"扣费交易金额"=订单额); 后面是天猫汇总列名
    amt_keys = ["扣费金额", "扣费金额(元)", "扣费金额合计 (元）", "扣费金额合计(元)",
                "金额", "支出金额合计（元）", "捐赠金额", "本月付款"]
    amt_col = None
    for k in amt_keys:
        if k in col:
            amt_col = col[k]
            break
    fee_type_col = col.get("业务大类", 1)

    def _num(v):
        try:
            return float(str(v).strip().rstrip("\t").strip())
        except (TypeError, ValueError):
            return 0.0

    out = []
    for r in all_rows[1:]:
        if not r or not r[0]:
            continue
        amt = _num(r[amt_col]) if (amt_col is not None and amt_col < len(r)) else 0.0
        ft = r[fee_type_col] if fee_type_col < len(r) else ""
        out.append({
            "fee_type": str(ft or "").strip().rstrip("\t").strip(),
            "amount": amt,
            "source": source_name,
        })
    return out


# ===== 天猫推广明细 csv (GBK 编码) =====
def parse_tmall_ads(buf: bytes) -> list[dict]:
    text = buf.decode("gbk", errors="ignore")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    header = rows[0]
    col = {h: i for i, h in enumerate(header)}
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        out.append({
            "date": r[col.get("日期", 0)],
            "channel_id": r[col.get("场景ID", 1)] if "场景ID" in col else "",
            "channel_name": r[col.get("场景名字", 2)] if "场景名字" in col else "",
            "spend": float(r[col.get("花费", -1)] or 0) if "花费" in col else 0,
            "sales": float(r[col.get("总成交金额", -1)] or 0) if "总成交金额" in col else 0,
            "orders": int(float(r[col.get("总成交笔数", -1)] or 0)) if "总成交笔数" in col else 0,
        })
    return out


# ===== 顺丰月结账单 xlsx (账单明细 sheet) =====
def parse_sf_logistics(buf: bytes, year_month: str) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    if "账单明细" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["账单明细"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    # r0=合并标题, r1=表头, r2+ 数据
    for r in rows[2:]:
        if not r[0]:
            continue
        # 顺丰日期格式 "04-01"，需结合 year_month 推年
        date_str = str(r[1]) if r[1] else ""
        if not date_str.startswith(year_month[5:]):  # "04" 月份过滤
            continue
        out.append({
            "carrier": "顺丰",
            "tracking": str(r[2] or "").strip(),
            "date": f"{year_month[:4]}-{date_str}",
            "from": r[3] or "",
            "to": r[4] or "",
            "weight": float(r[6] or 0),
            "amount": float(r[11] or 0),
            "discount": float(r[10] or 0),
            "service_type": r[7] or "",
            "operator": r[12] or "",
        })
    wb.close()
    return out


# ===== 中通月结账单 xlsx =====
def parse_zt_logistics(buf: bytes, year_month: str) -> list[dict]:
    """中通: Sheet1 主账单, 列: 账单日期/运单号/结算重量/金额/.../合计/目的省/目的市/运单发放/结算对象."""
    wb = openpyxl.load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h or "").strip() for h in next(rows)]
    except StopIteration:
        wb.close()
        return []

    def _idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return -1

    i_date = _idx("账单日期")
    i_track = _idx("运单号")
    i_wt = _idx("结算重量")
    i_amt = _idx("合计", "金额", "应结金额", "费用合计")  # 兼容有无"合计"列
    i_to = _idx("目的地市", "目的市")
    i_op = _idx("运单发放")
    out = []

    def g(r, i):
        return r[i] if 0 <= i < len(r) else None

    for r in rows:
        if i_date < 0 or not g(r, i_date):
            continue
        if not _is_apr(r[i_date], year_month):
            continue
        out.append({
            "carrier": "中通",
            "tracking": str(g(r, i_track) or "").strip(),
            "date": str(g(r, i_date))[:10] if g(r, i_date) else "",
            "from": "",
            "to": g(r, i_to) or "",
            "weight": float(g(r, i_wt) or 0),
            "amount": float(g(r, i_amt) or 0),
            "discount": 0,
            "service_type": "",
            "operator": g(r, i_op) or "",
        })
    wb.close()
    return out


# ===== 抖音订单 =====
def _strip_tab(v):
    """抖音字段值常带 \\t 前缀."""
    if v is None:
        return ""
    return str(v).strip().lstrip("\t").strip()


def parse_dy_orders(buf: bytes, year_month: str) -> tuple[list[dict], set]:
    """抖音订单 xlsx. 字段: 主订单/子订单/选购商品/商品ID/商家编码/商品数量/商品金额/
    订单提交时间/支付完成时间/订单状态/售后状态/订单类型/订单应付金额/运费/优惠/手续费/商家收入金额."""
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], set()
    header = all_rows[0]
    col = {h: i for i, h in enumerate(header)}
    out = []
    sku_set = set()
    for r in all_rows[1:]:
        if not r[0]:
            continue
        pay_t = _strip_tab(r[col.get("支付完成时间", -1)])
        submit_t = _strip_tab(r[col.get("订单提交时间", -1)])
        t = pay_t or submit_t
        if year_month not in t:
            continue
        sku = _strip_tab(r[col.get("商家编码", -1)])
        if sku:
            sku_set.add(sku)
        # v0.7: 从「快递信息」列提取运单号 (格式 "SF0220481613112-顺丰速运,商品..." / "76922..-中通快递,..")
        express = _strip_tab(r[col.get("快递信息", -1)])
        tracking = ""
        if express and express != "-":
            tracking = express.split(",")[0].split("-")[0].strip()
        out.append({
            "main_oid": _strip_tab(r[col.get("主订单编号", -1)]),
            "sub_oid": _strip_tab(r[col.get("子订单编号", -1)]),
            "create_t": submit_t,
            "pay_t": pay_t,
            "sku": sku,
            "title": r[col.get("选购商品", -1)] or "",
            "attr": "",
            "qty": r[col.get("商品数量", -1)] or 0,
            "price": r[col.get("商品金额", -1)] or 0,
            "paid": r[col.get("订单应付金额", -1)] or 0,  # 优惠后实付
            "tracking": tracking,  # v0.7 从快递信息列解析
            "status": r[col.get("订单状态", -1)] or "",
            "refund_status": r[col.get("售后状态", -1)] or "",
        })
    wb.close()
    return out, sku_set


def parse_dy_refunds(buf: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    for r in rows[1:]:
        if not r[0]:
            continue
        out.append({
            "refund_id": _strip_tab(r[col.get("售后单号", -1)]),
            "main_oid": _strip_tab(r[col.get("订单号", -1)]),
            "sku": _strip_tab(r[col.get("商家编码", -1)]),
            "title": r[col.get("商品名称", -1)] or "",
            "complete_t": str(r[col.get("商家退款时间", -1)] or r[col.get("售后完结时间", -1)] or ""),
            "amount": r[col.get("退商品金额（元）", -1)] or 0,
            "type": r[col.get("售后类型", -1)] or "",
            "reason": r[col.get("售后原因", -1)] or "",
            "to_buyer": r[col.get("退商品金额（元）", -1)] or 0,
            "to_platform": 0,
        })
    wb.close()
    return out


def parse_dy_platform_fee(buf: bytes, source_name: str, year_month: str = "") -> list[dict]:
    """抖音平台结算 csv. 只算"出账"且非资金转账类 + 按"动账时间" 过滤当月.
    排除场景: 提现/充值/转账/保证金 (商家自己的资金转移, 不是平台费)."""
    NON_FEE_SCENES = {"提现", "充值", "转账", "保证金", "保证金充值", "保证金退还"}
    text = buf.decode("utf-8-sig", errors="ignore")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    header = rows[0]
    col = {h: i for i, h in enumerate(header)}
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        direction = r[col.get("动账方向", -1)] if "动账方向" in col else ""
        if direction != "出账":
            continue
        scene = r[col.get("动账场景", -1)] or ""
        if scene in NON_FEE_SCENES:
            continue
        t = str(r[col.get("动账时间", -1)] or "") if "动账时间" in col else ""
        if year_month and year_month not in t:
            continue
        out.append({
            "fee_type": scene or "其他",
            "amount": float(r[col.get("动账金额", -1)] or 0),
            "source": source_name,
        })
    return out


# ===== 小红书 =====
def parse_xhs_orders(buf: bytes, year_month: str) -> tuple[list[dict], set]:
    """小红书千帆 75 列订单. 商家编码在 col 69 (按表头查)."""
    wb = openpyxl.load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], set()
    header = rows[0]
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    sku_set = set()
    for r in rows[1:]:
        if not r[0]:
            continue
        pay_t = str(r[col.get("支付时间", -1)] or "")
        create_t = str(r[col.get("订单创建时间", -1)] or "")
        t = pay_t or create_t
        if year_month not in t:
            continue
        sku = str(r[col.get("商家编码", -1)] or "").strip()
        if sku:
            sku_set.add(sku)
        out.append({
            "main_oid": r[col.get("订单号", -1)] or "",
            "sub_oid": "",
            "create_t": create_t,
            "pay_t": pay_t,
            "sku": sku,
            "title": r[col.get("商品名称", -1)] or "",
            "attr": r[col.get("SKU规格", -1)] or "",
            "qty": r[col.get("SKU件数", -1)] or 0,
            "price": r[col.get("商品总价(元)", -1)] or 0,
            "paid": r[col.get("商家应收金额(元)（支付金额）", -1)] or 0,
            "tracking": str(r[col.get("快递单号", -1)] or "").strip(),
            "status": r[col.get("订单状态", -1)] or "",
            "refund_status": r[col.get("售后状态", -1)] or "",
        })
    wb.close()
    return out, sku_set


def parse_xhs_refunds(buf: bytes) -> list[dict]:
    """小红书退款 — 注意无"商家编码"字段, sku 先留空, P5 后做 main_oid → sku join."""
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    for r in rows[1:]:
        if not r[0]:
            continue
        out.append({
            "refund_id": r[col.get("售后单号", -1)] or "",
            "main_oid": r[col.get("订单号", -1)] or "",
            "sku": "",  # 小红书退款表无商家编码
            "title": r[col.get("商品名称", -1)] or "",
            "complete_t": str(r[col.get("商家确认收货时间", -1)] or r[col.get("退货创建时间", -1)] or ""),
            "amount": r[col.get("申请售后金额(元)", -1)] or 0,
            "type": r[col.get("售后类型", -1)] or "",
            "reason": r[col.get("原因", -1)] or "",
            "to_buyer": r[col.get("申请售后金额(元)", -1)] or 0,
            "to_platform": 0,
        })
    wb.close()
    return out


def parse_xhs_platform_fee(buf: bytes, source_name: str) -> list[dict]:
    """小红书平台结算: 创建时间/交易类型描述/收入/支出/账户余额/业务单号. 只算支出列."""
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = rows[0]
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    for r in rows[1:]:
        if not r[0]:
            continue
        spend = r[col.get("支出（元）", -1)]
        if not spend:
            continue
        try:
            amount = float(spend)
        except (ValueError, TypeError):
            continue
        if amount <= 0:
            continue
        out.append({
            "fee_type": r[col.get("交易类型描述", -1)] or "其他",
            "amount": amount,
            "source": source_name,
        })
    wb.close()
    return out


# ===== 拼多多 =====
def parse_pdd_orders(buf: bytes, year_month: str) -> tuple[list[dict], set]:
    """拼多多订单 27 列. 商家编码-规格维度 col 16. 按 magic 读 .xlsx/.xls/.csv."""
    rows = _load_rows(buf)
    if not rows:
        return [], set()
    header = rows[0]
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    sku_set = set()
    for r in rows[1:]:
        if not r[0]:
            continue
        ship_t = r[col.get("发货时间", -1)]
        confirm_t = r[col.get("确认收货时间", -1)]
        deal_t = r[col.get("订单成交时间", -1)]
        # 优先按发货时间过滤,其次按成交
        t = str(ship_t or deal_t or confirm_t or "")
        if year_month not in t:
            continue
        sku = str(r[col.get("商家编码-规格维度", -1)] or "").strip().rstrip("\t").strip()
        if sku:
            sku_set.add(sku)
        # paid: 商家实收金额(已含平台补贴)
        paid_raw = r[col.get("商家实收金额(元)", -1)]
        try:
            paid = float(str(paid_raw).rstrip("\t").strip()) if paid_raw not in (None, "") else 0
        except (ValueError, TypeError):
            paid = 0
        qty_raw = r[col.get("商品数量(件)", -1)]
        try:
            qty = float(str(qty_raw).rstrip("\t").strip()) if qty_raw not in (None, "") else 0
        except (ValueError, TypeError):
            qty = 0
        price_raw = r[col.get("商品总价(元)", -1)]
        try:
            price = float(str(price_raw).rstrip("\t").strip()) if price_raw not in (None, "") else 0
        except (ValueError, TypeError):
            price = 0
        out.append({
            "main_oid": r[col.get("订单号", -1)] or "",
            "sub_oid": "",
            "create_t": str(deal_t or ""),
            "pay_t": str(ship_t or deal_t or ""),
            "sku": sku,
            "title": r[col.get("商品", -1)] or "",
            "attr": r[col.get("商品规格", -1)] or "",
            "qty": qty,
            "price": price,
            "paid": paid,
            "tracking": str(r[col.get("快递单号", -1)] or "").strip().rstrip("\t").strip(),
            "status": r[col.get("订单状态", -1)] or "",
            "refund_status": r[col.get("售后状态", -1)] or "",
        })
    return out, sku_set


def parse_pdd_platform_fee(buf: bytes, source_name: str) -> list[dict]:
    """拼多多收入支出: 真表头在 r4. 字段: 商户订单号/发生时间/收入金额(+元)/支出金额(-元)/账务类型/备注/业务描述.
    只取支出列 < 0 的行作为成本. 按 magic 读 .xlsx/.xls/.csv."""
    rows = _load_rows(buf)
    # 找真表头行 (含"商户订单号")
    header_idx = None
    for i, r in enumerate(rows[:15]):
        if r and r[0] == "商户订单号":
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    col = {h: i for i, h in enumerate(header) if h}
    i_spend = col.get("支出金额（-元）", -1)
    i_type = col.get("账务类型", -1)
    out = []

    def g(r, i):  # 边界安全取值 (csv 行可能比表头短)
        return r[i] if 0 <= i < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r or not r[0]:
            continue
        spend = g(r, i_spend)
        if not spend:
            continue
        try:
            amount = abs(float(str(spend).strip()))
        except (ValueError, TypeError):
            continue
        if amount <= 0:
            continue
        out.append({
            "fee_type": g(r, i_type) or "其他",
            "amount": amount,
            "source": source_name,
        })
    return out


def parse_pdd_refunds(buf: bytes) -> list[dict]:
    """拼多多退款明细. 列: 售后编号/订单编号/交易金额/售后状态/退款类型/退款金额/.../申请时间.
    只算 售后状态=退款成功(排除 已撤销/售后单失败); 按订单编号 join 到订单。"""
    rows = _load_rows(buf)
    if not rows:
        return []
    header = rows[0]
    col = {str(h).strip(): i for i, h in enumerate(header) if h not in (None, "")}
    i_oid = col.get("订单编号", -1)
    i_amt = col.get("退款金额", -1)
    i_status = col.get("售后状态", -1)
    i_type = col.get("退款类型", -1)
    i_rid = col.get("售后编号", -1)
    i_apply = col.get("申请时间", -1)
    out = []

    def g(r, i):
        return r[i] if 0 <= i < len(r) else None

    for r in rows[1:]:
        if not r or g(r, i_oid) in (None, ""):
            continue
        if str(g(r, i_status) or "").strip() != "退款成功":
            continue  # 排除 已撤销/售后单失败
        try:
            amt = float(str(g(r, i_amt) or 0).strip())
        except (ValueError, TypeError):
            amt = 0
        if amt <= 0:
            continue
        out.append({
            "refund_id": str(g(r, i_rid) or ""),
            "main_oid": str(g(r, i_oid) or "").strip(),
            "sku": "", "title": "",
            "complete_t": str(g(r, i_apply) or ""),
            "amount": amt,
            "type": str(g(r, i_type) or ""), "reason": "",
            "to_buyer": amt, "to_platform": 0,
        })
    return out


# ===== 京东 (v0.3 完整版) =====
# 京东"货款明细" 是 30 列费用流水(同 1 订单多行=多种费用项):
# 字段: 订单编号/订单状态/订单下单时间/订单完成时间/商品编号/商品名称/商品数量/
#       扣点类型/佣金比例/费用名称/应结金额/收支方向(收入/支出)/...
# v0.3 策略: 按"订单编号"聚合 → 收入合计=销售额, 支出合计=平台费.
# 京东"商品编号"是京东ID不是商家编码 → SKU 标"未填" (待 v0.4 补 京东商品 ↔ ERP_SKU 映射)
def _strip_jd(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.startswith("'"):
        s = s[1:]
    return s.rstrip("\t").strip()


def _parse_jd_flow(buf: bytes, ext: str) -> tuple[list[tuple], list[str]]:
    """读 30 列费用流水, 返回 (rows, header). 支持 csv (utf-8-sig/gbk) 和 xlsx."""
    if ext == "csv":
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                rows = list(csv.reader(io.StringIO(buf.decode(enc))))
                if rows and "订单编号" in (rows[0][0] or ""):
                    return rows[1:], rows[0]
            except Exception:
                continue
        return [], []
    else:
        wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []
        return list(all_rows[1:]), list(all_rows[0])


def _read_jd_table(buf: bytes, ext: str) -> tuple[list, list]:
    """通用读 (rows, header), 不假设格式. 支持 csv (utf-8-sig/gbk) 与 xlsx."""
    if ext == "csv":
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                rows = list(csv.reader(io.StringIO(buf.decode(enc))))
                if rows:
                    return rows[1:], rows[0]
            except Exception:
                continue
        return [], []
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return [], []
    return list(all_rows[1:]), list(all_rows[0])


# 京东订单状态: 这些视为未成交, 不计销售 (退款另由退款明细处理)
JD_INVALID_STATUS = ("待付款", "待支付", "等待付款", "已取消", "取消", "已关闭", "关闭")


def parse_jd_orders(buf: bytes, year_month: str, ext: str = "xlsx") -> tuple[list[dict], set]:
    """京东订单 (赵伟俊 2026-06 口径): 读标准「订单明细」导出。
    销售额=应付金额(=货款) / SKU=商家SKUID(干净ERP SKU) / 数量=订购数量 / 运单=快递单号 / 归月=下单时间。
    平台费另由「订单结算明细对账表」走 parse_jd_platform_fee(支出列)。退款另由退款明细 join。
    旧版「货款明细流水」格式(订单编号+收支方向)不再作订单源 → 返回空。"""
    rows, header = _read_jd_table(buf, ext)
    if not header:
        return [], set()
    col = {str(h).strip(): i for i, h in enumerate(header) if h}
    if "订单号" not in col or "应付金额" not in col:
        return [], set()  # 非订单明细格式(如对账流水) 不作订单源

    def g(r, k):
        i = col.get(k, -1)
        return r[i] if 0 <= i < len(r) else None

    out = []
    sku_set = set()
    for r in rows:
        if not r or not r[0]:
            continue
        oid = _strip_jd(g(r, "订单号"))
        if not oid:
            continue
        status = _strip_jd(g(r, "订单状态"))
        if any(s in status for s in JD_INVALID_STATUS):
            continue
        order_t = _strip_jd(g(r, "下单时间"))
        if year_month not in (order_t or ""):
            continue  # 归月按下单时间, 与其他平台对齐
        try:
            paid = float(_strip_jd(g(r, "应付金额")) or 0)
        except (ValueError, TypeError):
            paid = 0
        if paid <= 0:
            continue
        sku = _strip_jd(g(r, "商家SKUID"))
        try:
            qty = float(_strip_jd(g(r, "订购数量")) or 1)
        except (ValueError, TypeError):
            qty = 1
        if sku:
            sku_set.add(sku)
        out.append({
            "main_oid": oid,
            "sub_oid": "",
            "create_t": order_t,
            "pay_t": _strip_jd(g(r, "付款确认时间")),
            "sku": sku or "(未填)",
            "title": _strip_jd(g(r, "商品名称")) or "",
            "attr": "",
            "qty": qty or 1,
            "price": 0,
            "paid": paid,  # 应付金额 = 货款 = 销售额 (赵口径)
            "tracking": _strip_jd(g(r, "快递单号")),
            "status": status,
            "refund_status": "",
        })
    return out, sku_set


def parse_jd_refunds(buf: bytes) -> list[dict]:
    """京东退款明细 47 列, 含订单号/商品编号/退款金额. 用作退款源 + 反推销售."""
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return []
    # r0 是分组标题(申请信息/订单信息/...), r1 是真实表头, r2 是数据
    if rows[0] and rows[0][0] == "申请信息":
        header = rows[1]
        data_start = 2
    else:
        header = rows[0]
        data_start = 1
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    for r in rows[data_start:]:
        if not r or not r[0]:
            continue
        out.append({
            "refund_id": r[col.get("服务单号", -1)] or "",
            "main_oid": str(r[col.get("订单号", -1)] or ""),
            "sku": "",  # 商品编号是京东ID不是商家编码, v0.3 补 mapping
            "title": r[col.get("商品名称", -1)] or "",
            "complete_t": str(r[col.get("商家首次处理时间", -1)] or r[col.get("审核时间", -1)] or ""),
            "amount": r[col.get("退款金额", -1)] or 0,
            "type": r[col.get("客户期望", -1)] or "",
            "reason": r[col.get("一级申请原因", -1)] or "",
            "to_buyer": r[col.get("退款金额", -1)] or 0,
            "to_platform": 0,
        })
    wb.close()
    return out


def parse_jd_platform_fee(buf: bytes, source_name: str, ext: str = "csv",
                          year_month: str = "") -> list[dict]:
    """京东平台费. 流水(30列): 按月+排除货款/提现/保证金. 汇总csv: 跳过(避免与流水重复算)."""
    if ext == "csv":
        text = None
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = buf.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            return []
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            return []
        if rows[0] and "订单编号" in (rows[0][0] or ""):
            return _jd_flow_to_fees(rows[1:], rows[0], source_name, year_month)
        # 汇总 csv 跳过 (避免与货款明细重复). 京东必须上传货款明细才能算平台费.
        return []
    else:
        wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return []
        return _jd_flow_to_fees(list(all_rows[1:]), list(all_rows[0]), source_name, year_month)


# 京东"支出"中的非平台费类型 (资金转移/货款本体)
JD_NON_FEE_NAMES = {"货款", "提现", "保证金", "保证金充值", "保证金退还", "充值", "转账"}


def _jd_flow_to_fees(rows, header, source_name, year_month=""):
    col = {h: i for i, h in enumerate(header) if h}
    out = []
    for r in rows:
        if not r or not r[0]:
            continue
        direction = _strip_jd(r[col.get("收支方向", -1)])
        if direction != "支出":
            continue
        fee_name = _strip_jd(r[col.get("费用名称", -1)])
        if fee_name in JD_NON_FEE_NAMES:
            continue
        if year_month:
            t = _strip_jd(r[col.get("到账时间", -1)])
            if not t:
                t = _strip_jd(r[col.get("账单生成时间", -1)])
            ym_compact = year_month.replace("-", "")
            if year_month not in t and ym_compact not in t:
                continue
        amt_raw = _strip_jd(r[col.get("应结金额", -1)])
        try:
            amt = abs(float(amt_raw))
        except (ValueError, TypeError):
            continue
        if amt <= 0:
            continue
        out.append({
            "fee_type": fee_name or "其他",
            "amount": amt,
            "source": source_name,
        })
    return out


# ===== 解析分发器 (按平台 + kind 路由) =====
def detect_and_parse(filename: str, buf: bytes, year_month: str, kind_hint: str,
                     platform: str = "天猫") -> dict:
    """按 platform + kind_hint 路由. platform: 天猫/抖音/小红书/京东.
    返回 {"kind": ..., "data": [...]} 或 {"kind": "error", "msg": ...}."""
    fn = filename.lower()
    try:
        # 物流不分平台
        if kind_hint == "物流":
            if "顺丰" in filename or "sf" in fn:
                return {"kind": "物流", "data": parse_sf_logistics(buf, year_month)}
            if "中通" in filename or "zt" in fn:
                return {"kind": "物流", "data": parse_zt_logistics(buf, year_month)}
            return {"kind": "error", "msg": f"未识别快递公司: {filename}"}

        # 平台特定 parser
        if platform in ("天猫", "淘宝"):  # 淘宝结构与天猫一致, 复用同一组 parser
            if kind_hint == "订单":
                data, skus = parse_tmall_orders(buf, year_month, filename)
                return {"kind": "订单", "data": data, "sku_set": list(skus)}
            if kind_hint == "退款":
                return {"kind": "退款", "data": parse_tmall_refunds(buf, filename)}
            if kind_hint == "平台费":
                return {"kind": "平台费", "data": parse_tmall_platform_fee(buf, filename, filename)}
            if kind_hint == "广告":
                return {"kind": "广告", "data": parse_tmall_ads(buf)}
        elif platform == "抖音":
            if kind_hint == "订单":
                data, skus = parse_dy_orders(buf, year_month)
                return {"kind": "订单", "data": data, "sku_set": list(skus)}
            if kind_hint == "退款":
                return {"kind": "退款", "data": parse_dy_refunds(buf)}
            if kind_hint == "平台费":
                return {"kind": "平台费", "data": parse_dy_platform_fee(buf, filename, year_month)}
            if kind_hint == "广告":
                return {"kind": "广告", "data": []}
        elif platform == "小红书":
            if kind_hint == "订单":
                data, skus = parse_xhs_orders(buf, year_month)
                return {"kind": "订单", "data": data, "sku_set": list(skus)}
            if kind_hint == "退款":
                return {"kind": "退款", "data": parse_xhs_refunds(buf)}
            if kind_hint == "平台费":
                return {"kind": "平台费", "data": parse_xhs_platform_fee(buf, filename)}
            if kind_hint == "广告":
                return {"kind": "广告", "data": []}
        elif platform == "拼多多":
            if kind_hint == "订单":
                data, skus = parse_pdd_orders(buf, year_month)
                return {"kind": "订单", "data": data, "sku_set": list(skus)}
            if kind_hint == "退款":
                return {"kind": "退款", "data": parse_pdd_refunds(buf)}
            if kind_hint == "平台费":
                return {"kind": "平台费", "data": parse_pdd_platform_fee(buf, filename)}
            if kind_hint == "广告":
                return {"kind": "广告", "data": []}
        elif platform == "京东":
            ext = "csv" if fn.endswith(".csv") else "xlsx"
            if kind_hint == "订单":
                data, skus = parse_jd_orders(buf, year_month, ext)
                return {"kind": "订单", "data": data, "sku_set": list(skus)}
            if kind_hint == "退款":
                return {"kind": "退款", "data": parse_jd_refunds(buf)}
            if kind_hint == "平台费":
                return {"kind": "平台费",
                        "data": parse_jd_platform_fee(buf, filename, ext, year_month)}
            if kind_hint == "广告":
                return {"kind": "广告", "data": []}
        return {"kind": "error", "msg": f"v0.2.7 暂不支持 {platform}/{kind_hint}"}
    except Exception as e:
        return {"kind": "error", "msg": f"解析 {filename} ({platform}/{kind_hint}): {type(e).__name__}: {e}"}
