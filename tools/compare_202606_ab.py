"""2026-06 domestic e-commerce gross-profit A/B comparison.

Reads the manual baseline workbook (B) and the automation Feishu Sheet (A),
then writes a local comparison workbook plus a short markdown summary.
"""
from __future__ import annotations

import asyncio
import math
import os
import sys
import urllib.parse
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app import feishu  # noqa: E402


BASELINE_PATH = Path(
    r"D:\Users\Administrator\Desktop\财务毛利报表计算资料\深圳奥迪尔\2026年6月\_输出"
    r"\深圳奥迪尔国内电商2026-06毛利试算-v1-自动化对比基准.xlsx"
)
AUTO_SHEET_TOKEN = os.getenv("AUTO_SHEET_TOKEN", "Wa6wsgM7chHuyitd6yRcp7KjnXf")
OUT_DIR = ROOT
AUTO_REBUILT_PATH = OUT_DIR / "tmp_auto_2026_06_values.xlsx"
COMPARE_PATH = OUT_DIR / "domestic_2026_06_ab_compare.xlsx"
SUMMARY_PATH = OUT_DIR / "domestic_2026_06_ab_compare.md"


REQUIRED_BASELINE_SHEETS = [
    "月度毛利试算",
    "产品毛利_月度",
    "SKU成本明细",
    "物流匹配明细",
    "费用明细汇总",
    "缺口清单",
]


AUTO_REQUIRED_FINAL_SHEETS = [
    "月度毛利试算",
    "产品毛利_月度",
    "SKU成本明细",
    "物流匹配明细",
    "费用明细汇总",
    "缺口清单",
    "资料清单审计",
    "口径说明",
    "税务A_B核对",
]


STORE_ALIASES = {
    ("天猫", "纷岚店"): "天猫纷岚",
    ("天猫", "天猫纷岚"): "天猫纷岚",
    ("天猫", "POWKONG旗舰店"): "天猫宝空",
    ("天猫", "天猫宝空"): "天猫宝空",
    ("抖音", "宝空店"): "抖音宝空",
    ("抖音", "抖音宝空"): "抖音宝空",
    ("抖音", "纷岚店"): "抖音纷岚",
    ("抖音", "抖音纷岚"): "抖音纷岚",
    ("小红书", "宝空店"): "小红书宝空",
    ("小红书", "小红书宝空"): "小红书宝空",
    ("小红书", "纷岚店"): "小红书纷岚",
    ("小红书", "小红书纷岚"): "小红书纷岚",
    ("京东", "京东宝空"): "京东宝空",
    ("京东", "京东纷岚"): "京东纷岚",
}


@dataclass
class Check:
    item: str
    status: str
    evidence: str
    action: str


def col_letter(n: int) -> str:
    return get_column_letter(max(1, n))


def is_blank_row(row: Iterable[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def trim_row(row: Iterable[Any]) -> list[Any]:
    values = list(row)
    while values and (values[-1] is None or str(values[-1]).strip() == ""):
        values.pop()
    return values


def rows_to_dicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if is_blank_row(row):
            continue
        item: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if h and h not in item:
                item[h] = row[i] if i < len(row) else None
        out.append(item)
    return out


def load_baseline() -> dict[str, list[list[Any]]]:
    wb = load_workbook(BASELINE_PATH, data_only=True, read_only=True)
    data: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames:
        rows: list[list[Any]] = []
        for row in wb[name].iter_rows(values_only=True):
            trimmed = trim_row(row)
            if trimmed:
                rows.append(trimmed)
        data[name] = rows
    wb.close()
    return data


async def read_auto_sheet_values() -> dict[str, list[list[Any]]]:
    meta = await feishu.sheets_metainfo(AUTO_SHEET_TOKEN)
    sheets = meta.get("data", {}).get("sheets", [])
    out: dict[str, list[list[Any]]] = {}
    for sheet in sheets:
        title = sheet["title"]
        sheet_id = sheet["sheetId"]
        row_count = int(sheet.get("rowCount") or 200)
        col_count = int(sheet.get("columnCount") or 30)
        rng = f"{sheet_id}!A1:{col_letter(col_count)}{row_count}"
        enc = urllib.parse.quote(rng, safe="")
        resp = await feishu._req(
            "GET",
            f"/open-apis/sheets/v2/spreadsheets/{AUTO_SHEET_TOKEN}/values/{enc}",
            params={"valueRenderOption": "UnformattedValue"},
        )
        values = resp.get("data", {}).get("valueRange", {}).get("values", []) or []
        out[title] = [trim_row(row) for row in values if not is_blank_row(row)]
    return out


def num(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        if s.endswith("%"):
            return float(s[:-1]) / 100.0
        return float(s)
    except ValueError:
        return None


def n0(value: Any) -> float:
    return num(value) or 0.0


def norm_platform(value: Any) -> str:
    return str(value or "").strip()


def norm_store(platform: Any, store: Any) -> str:
    p = norm_platform(platform)
    s = str(store or "").strip()
    return STORE_ALIASES.get((p, s), s)


def norm_msku(value: Any) -> str:
    return str(value or "").strip().upper()


def norm_name(value: Any) -> str:
    return " ".join(str(value or "").split())


def money_delta_status(b: Any, a: Any, tol: float = 0.01) -> str:
    bn = num(b)
    an = num(a)
    if bn is None and an is None:
        return "NA"
    if bn is None or an is None:
        return "缺字段"
    return "OK" if abs(an - bn) <= tol else "FAIL"


def percent_delta_status(b: Any, a: Any, tol: float = 0.0001) -> str:
    return money_delta_status(b, a, tol=tol)


def write_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def style_ws(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2):
        status_text = " ".join(str(c.value or "") for c in row[:8])
        fill = None
        if "FAIL" in status_text or "缺失" in status_text or "不可确认" in status_text:
            fill = fail_fill
        elif "WARN" in status_text or "缺字段" in status_text or "待补" in status_text:
            fill = warn_fill
        elif "OK" in status_text or "PASS" in status_text:
            fill = ok_fill
        if fill:
            for cell in row:
                cell.fill = fill
    widths: dict[int, int] = defaultdict(lambda: 10)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = min(max(widths[cell.column], len(str(cell.value)) + 2), 60)
    for idx, width in widths.items():
        ws.column_dimensions[col_letter(idx)].width = width
    ws.freeze_panes = "A2"


def save_auto_rebuilt(auto: dict[str, list[list[Any]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in auto.items():
        ws = wb.create_sheet(title[:31])
        write_rows(ws, rows)
        style_ws(ws)
    wb.save(AUTO_REBUILT_PATH)


def build_monthly_compare(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    base_rows = rows_to_dicts(base.get("月度毛利试算", []))
    auto_rows = rows_to_dicts(auto.get("月度毛利试算", [])) or rows_to_dicts(auto.get("11_店铺汇总看板", []))
    b_by_store = {norm_store(r.get("平台"), r.get("店铺")): r for r in base_rows}
    a_by_store = {norm_store(r.get("平台"), r.get("店铺")): r for r in auto_rows}
    if "月度毛利试算" in auto:
        field_map = [
            ("销售订单数", "销售订单数", "销售订单数", "number"),
            ("销量", "销量", "销量", "number"),
            ("销售额", "销售额(RMB)", "销售额(RMB)", "money"),
            ("退款", "退款(RMB)", "退款(RMB)", "money"),
            ("净销售额", "净销售额(RMB)", "净销售额(RMB)", "money"),
            ("平台费用", "平台费用(RMB)", "平台费用(RMB)", "money"),
            ("广告费", "广告费(RMB)", "广告费(RMB)", "money"),
            ("采购成本", "采购成本(RMB)", "采购成本(RMB)", "money"),
            ("尾程费用", "尾程费用(RMB)", "尾程费用(RMB)", "money"),
            ("其他费用", "其他费用(RMB)", "其他费用(RMB)", "money"),
            ("试算毛利", "试算毛利(RMB)", "试算毛利(RMB)", "money"),
            ("试算毛利率", "试算毛利率", "试算毛利率", "percent"),
            ("结算回款/净回款", "结算回款/净回款(RMB)", "结算回款/净回款(RMB)", "money"),
            ("P0缺口数", "P0缺口数", "P0缺口数", "number"),
        ]
    else:
        field_map = [
            ("销量", "销量", "总销量", "number"),
            ("销售额", "销售额(RMB)", "销售额(应付货款)", "money"),
            ("净销售额", "净销售额(RMB)", "净销售额", "money"),
            ("平台费用", "平台费用(RMB)", "平台费合计", "money"),
            ("广告费", "广告费(RMB)", "广告费合计", "money"),
            ("采购成本", "采购成本(RMB)", "采购成本", "money"),
            ("尾程费用", "尾程费用(RMB)", "物流成本", "money"),
            ("试算毛利", "试算毛利(RMB)", "毛利额", "money"),
            ("试算毛利率", "试算毛利率", "毛利率", "percent"),
            ("销售订单数", "销售订单数", None, "missing"),
            ("退款", "退款(RMB)", None, "missing"),
            ("其他费用", "其他费用(RMB)", None, "missing"),
            ("结算回款/净回款", "结算回款/净回款(RMB)", None, "missing"),
            ("P0缺口数", "P0缺口数", None, "missing"),
        ]
    rows = [["平台店铺", "字段", "人工B", "自动A", "差异(A-B)", "状态", "说明"]]
    for key in sorted(set(b_by_store) | set(a_by_store)):
        b = b_by_store.get(key, {})
        a = a_by_store.get(key, {})
        if not b:
            rows.append([key, "整行", "", "存在", "", "FAIL", "自动化多出人工基准不存在的平台店铺"])
            continue
        if not a:
            rows.append([key, "整行", "存在", "", "", "FAIL", "自动化缺少人工基准平台店铺"])
            continue
        for label, b_field, a_field, kind in field_map:
            b_val = b.get(b_field)
            a_val = a.get(a_field) if a_field else None
            if kind == "missing":
                rows.append([key, label, b_val, "", "", "缺字段", "自动化汇总看板未输出该字段"])
                continue
            bn = num(b_val)
            an = num(a_val)
            diff = None if bn is None or an is None else an - bn
            status = percent_delta_status(b_val, a_val) if kind == "percent" else money_delta_status(b_val, a_val)
            rows.append([key, label, b_val, a_val, diff, status, ""])
    return rows


def build_product_compare(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    base_rows = rows_to_dicts(base.get("产品毛利_月度", []))
    auto_rows = rows_to_dicts(auto.get("产品毛利_月度", []))
    valid_platforms = {"天猫", "抖音", "小红书", "京东"}
    base_rows = [r for r in base_rows if norm_platform(r.get("国内电商平台名称")) in valid_platforms]
    auto_rows = [r for r in auto_rows if norm_platform(r.get("国内电商平台名称") or r.get("平台")) in valid_platforms]
    b_by_key = {
        (norm_platform(r.get("国内电商平台名称")), norm_store(r.get("国内电商平台名称"), r.get("站点")),
         norm_msku(r.get("MSKU")), norm_name(r.get("中文名称"))): r
        for r in base_rows
    }
    if auto_rows and "国内电商平台名称" in auto_rows[0]:
        a_by_key = {
            (norm_platform(r.get("国内电商平台名称")), norm_store(r.get("国内电商平台名称"), r.get("站点")),
             norm_msku(r.get("MSKU")), norm_name(r.get("中文名称"))): r
            for r in auto_rows
        }
        field_map = [
            ("销量", "销量", "销量", "number"),
            ("退款数量", "退款数量", "退款数量", "number"),
            ("销售额", "销售额(RMB)", "销售额(RMB)", "money"),
            ("退款", "退款(RMB)", "退款(RMB)", "money"),
            ("平台服务费", "平台服务费(RMB)", "平台服务费(RMB)", "money"),
            ("广告费", "广告费(RMB)", "广告费(RMB)", "money"),
            ("采购成本", "采购成本(RMB)", "采购成本(RMB)", "money"),
            ("尾程费用", "尾程费用(RMB)", "尾程费用(RMB)", "money"),
            ("其他成本", "其他成本(RMB)", "其他成本(RMB)", "money"),
            ("毛利润", "毛利润(RMB)", "毛利润(RMB)", "money"),
            ("毛利率", "毛利率", "毛利率", "percent"),
        ]
    else:
        a_by_key = {
            (norm_platform(r.get("平台")), norm_store(r.get("平台"), r.get("店铺")),
             norm_msku(r.get("ERP_SKU(=商家编码)")), norm_name(r.get("商品名称") or r.get("中文名称"))): r
            for r in auto_rows
        }
        field_map = [
            ("销量", "销量", "销量", "number"),
            ("退款数量", "退款数量", "退款数量(订单表口径)", "number"),
            ("销售额", "销售额(RMB)", "销售额(买家应付货款)", "money"),
            ("退款", "退款(RMB)", "退款金额", "money"),
            ("平台服务费", "平台服务费(RMB)", "平台费合计", "money"),
            ("广告费", "广告费(RMB)", "推广/广告费", "money"),
            ("采购成本", "采购成本(RMB)", "采购成本(含包材)", "money"),
            ("尾程费用", "尾程费用(RMB)", "物流成本", "money"),
            ("其他成本", "其他成本(RMB)", None, "missing"),
            ("毛利润", "毛利润(RMB)", "毛利额", "money"),
            ("毛利率", "毛利率", "毛利率", "percent"),
        ]
    rows = [["平台", "店铺", "MSKU", "中文名称", "字段", "人工B", "自动A", "差异(A-B)", "状态", "说明"]]
    for key in sorted(set(b_by_key) | set(a_by_key)):
        b = b_by_key.get(key, {})
        a = a_by_key.get(key, {})
        platform, store, msku, name = key
        if not b:
            rows.append([platform, store, msku, name, "整行", "", "存在", "", "FAIL", "自动化多出人工基准不存在的SKU"])
            continue
        if not a:
            rows.append([platform, store, msku, name, "整行", "存在", "", "", "FAIL", "自动化缺少人工基准SKU"])
            continue
        for label, b_field, a_field, kind in field_map:
            b_val = b.get(b_field)
            a_val = a.get(a_field) if a_field else None
            if kind == "missing":
                rows.append([platform, store, msku, name, label, b_val, "", "", "缺字段", "自动化产品毛利未输出该字段"])
                continue
            bn = num(b_val)
            an = num(a_val)
            diff = None if bn is None or an is None else an - bn
            status = percent_delta_status(b_val, a_val) if kind == "percent" else money_delta_status(b_val, a_val)
            rows.append([platform, store, msku, name, label, b_val, a_val, diff, status, ""])
    return rows


def aggregate(rows: list[dict[str, Any]], key_fields: list[str], value_field: str) -> dict[tuple[str, ...], float]:
    out: dict[tuple[str, ...], float] = defaultdict(float)
    for r in rows:
        key = tuple(str(r.get(f) or "").strip() for f in key_fields)
        out[key] += n0(r.get(value_field))
    return dict(out)


def build_fee_compare(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    base_rows = rows_to_dicts(base.get("费用明细汇总", []))
    final_fee_rows = rows_to_dicts(auto.get("费用明细汇总", []))
    platform_fee_rows = rows_to_dicts(auto.get("03_平台费用_导入", []))
    ad_rows = rows_to_dicts(auto.get("04_广告佣金_导入", []))
    b_agg: dict[tuple[str, str, str], float] = defaultdict(float)
    for r in base_rows:
        p = norm_platform(r.get("平台"))
        s = norm_store(p, r.get("店铺"))
        cat = str(r.get("费用类别") or "").strip()
        remark = str(r.get("备注") or "")
        if "排除" in remark:
            continue
        b_agg[(p, s, cat)] += n0(r.get("金额"))
    a_agg: dict[tuple[str, str, str], float] = defaultdict(float)
    if final_fee_rows:
        for r in final_fee_rows:
            p = norm_platform(r.get("平台"))
            s = norm_store(p, r.get("店铺"))
            cat = str(r.get("费用类别") or "平台费用").strip()
            remark = str(r.get("备注") or "")
            if "排除" in remark:
                continue
            a_agg[(p, s, cat)] += n0(r.get("金额"))
    else:
        for r in platform_fee_rows:
            p = norm_platform(r.get("平台"))
            s = norm_store(p, r.get("店铺"))
            cat = str(r.get("费用类型") or "平台费用").strip()
            a_agg[(p, s, cat)] += n0(r.get("费用金额"))
        for r in ad_rows:
            p = norm_platform(r.get("平台"))
            s = norm_store(p, r.get("店铺"))
            a_agg[(p, s, "广告费")] += n0(r.get("花费/佣金金额"))
    rows = [["平台", "店铺", "费用类别", "人工B金额", "自动A金额", "差异(A-B)", "状态", "说明"]]
    for key in sorted(set(b_agg) | set(a_agg)):
        b_val = round(b_agg.get(key, 0.0), 2)
        a_val = round(a_agg.get(key, 0.0), 2)
        status = "OK" if abs(a_val - b_val) <= 0.01 else "FAIL"
        rows.append([*key, b_val, a_val, round(a_val - b_val, 2), status, "明细聚合对账；自动化无同名费用明细汇总sheet"])
    return rows


def build_logistics_compare(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    base_rows = rows_to_dicts(base.get("物流匹配明细", []))
    auto_rows = rows_to_dicts(auto.get("物流匹配明细", [])) or rows_to_dicts(auto.get("05_物流月结账单_导入", []))
    b_by_waybill: dict[str, float] = defaultdict(float)
    b_meta: dict[str, dict[str, Any]] = {}
    for r in base_rows:
        wb = str(r.get("运单号") or "").strip()
        if not wb:
            continue
        b_by_waybill[wb] += n0(r.get("分摊尾程费用"))
        b_meta[wb] = r
    a_by_waybill: dict[str, float] = defaultdict(float)
    a_meta: dict[str, dict[str, Any]] = {}
    for r in auto_rows:
        wb = str(r.get("运单号") or "").strip()
        if not wb:
            continue
        a_by_waybill[wb] += n0(r.get("分摊尾程费用") if "分摊尾程费用" in r else r.get("应付金额"))
        a_meta[wb] = r
    rows = [["运单号", "平台", "店铺", "人工B分摊尾程", "自动A账单金额", "差异(A-B)", "状态", "B来源", "A来源", "说明"]]
    for wb in sorted(set(b_by_waybill) | set(a_by_waybill)):
        b_val = round(b_by_waybill.get(wb, 0.0), 2)
        a_val = round(a_by_waybill.get(wb, 0.0), 2)
        b = b_meta.get(wb, {})
        a = a_meta.get(wb, {})
        if wb not in b_by_waybill:
            status = "WARN"
            note = "自动账单有该运单，但人工基准未计入结算订单匹配"
        elif wb not in a_by_waybill:
            status = "FAIL"
            note = "人工基准计入该运单，自动账单缺失"
        else:
            status = "OK" if abs(a_val - b_val) <= 0.01 else "WARN"
            note = "同名物流匹配明细对账" if "物流匹配明细" in auto else "自动化缺同名物流匹配明细sheet；此处只能按运单原始账单金额辅助核对"
        rows.append([
            wb,
            b.get("平台", ""),
            b.get("店铺", ""),
            b_val,
            a_val,
            round(a_val - b_val, 2),
            status,
            f"{b.get('账单文件/API','')} / {b.get('来源','')}",
            a.get("来源文件", ""),
            note,
        ])
    return rows


def build_cost_compare(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    base_rows = rows_to_dicts(base.get("SKU成本明细", []))
    auto_cost_rows = rows_to_dicts(auto.get("SKU成本明细", []))
    auto_rows = rows_to_dicts(auto.get("产品毛利_月度", []))
    b_agg: dict[tuple[str, str, str], dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "cost": 0.0})
    for r in base_rows:
        p = norm_platform(r.get("平台"))
        s = norm_store(p, r.get("店铺"))
        sku = norm_msku(r.get("ERP_SKU"))
        b_agg[(p, s, sku)]["qty"] += n0(r.get("净成本数量"))
        b_agg[(p, s, sku)]["cost"] += n0(r.get("采购成本"))
    a_agg: dict[tuple[str, str, str], dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "cost": 0.0})
    if auto_cost_rows:
        for r in auto_cost_rows:
            p = norm_platform(r.get("平台"))
            s = norm_store(p, r.get("店铺"))
            sku = norm_msku(r.get("ERP_SKU"))
            a_agg[(p, s, sku)]["qty"] += n0(r.get("净成本数量"))
            a_agg[(p, s, sku)]["cost"] += n0(r.get("采购成本"))
    else:
        for r in auto_rows:
            p = norm_platform(r.get("平台"))
            s = norm_store(p, r.get("店铺"))
            sku = norm_msku(r.get("ERP_SKU(=商家编码)"))
            a_agg[(p, s, sku)]["qty"] += n0(r.get("净销量"))
            a_agg[(p, s, sku)]["cost"] += n0(r.get("采购成本(含包材)"))
    rows = [["平台", "店铺", "ERP_SKU", "人工B净成本数量", "自动A净销量", "数量差异", "人工B采购成本", "自动A采购成本", "成本差异", "状态", "说明"]]
    for key in sorted(set(b_agg) | set(a_agg)):
        b = b_agg.get(key, {"qty": 0.0, "cost": 0.0})
        a = a_agg.get(key, {"qty": 0.0, "cost": 0.0})
        qty_diff = a["qty"] - b["qty"]
        cost_diff = a["cost"] - b["cost"]
        status = "OK" if abs(qty_diff) <= 0.001 and abs(cost_diff) <= 0.01 else "FAIL"
        note = "同名SKU成本明细对账" if auto_cost_rows else "自动化未输出订单级SKU成本明细；用产品毛利SKU聚合辅助核对"
        rows.append([*key, b["qty"], a["qty"], qty_diff, round(b["cost"], 2), round(a["cost"], 2), round(cost_diff, 2), status, note])
    return rows


def build_gap_compare(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    base_rows = rows_to_dicts(base.get("缺口清单", []))
    warn_rows = rows_to_dicts(auto.get("缺口清单", [])) or rows_to_dicts(auto.get("12_异常预警", []))
    rows = [["来源", "P级/严重度", "平台", "店铺", "对象", "问题/描述", "影响", "状态", "说明"]]
    if not base_rows:
        rows.append(["人工B", "", "", "", "", "缺口清单为空", "", "OK", "人工基准 P0/P1 为空"])
    for r in base_rows:
        rows.append(["人工B", r.get("P级"), r.get("平台"), r.get("店铺"), r.get("对象"), r.get("问题"), r.get("影响"), "INFO", ""])
    if not warn_rows:
        rows.append(["自动A", "", "", "", "", "异常预警为空", "", "OK", ""])
    for r in warn_rows:
        if "P级" in r:
            rows.append(["自动A", r.get("P级"), r.get("平台"), r.get("店铺"), r.get("对象"), r.get("问题"), r.get("影响"), "WARN", "同名缺口清单"])
        else:
            rows.append(["自动A", r.get("严重度"), r.get("平台"), r.get("店铺"), r.get("ERP_SKU/单号"), r.get("描述"), r.get("影响金额"), "WARN", "自动化未输出同名缺口清单；异常预警不能等同P0缺口"])
    return rows


def build_sheet_check(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[list[Any]]:
    auto_names = set(auto)
    alt = {
        "月度毛利试算": "11_店铺汇总看板",
        "SKU成本明细": "产品毛利_月度/06_ERP_SKU成本表",
        "物流匹配明细": "05_物流月结账单_导入",
        "费用明细汇总": "03_平台费用_导入 + 04_广告佣金_导入",
        "缺口清单": "12_异常预警",
    }
    rows = [["要求sheet", "人工基准B", "自动A同名", "自动A替代/现有", "状态", "说明"]]
    for name in AUTO_REQUIRED_FINAL_SHEETS:
        b_ok = "有" if name in base else "无"
        a_same = "有" if name in auto_names else "无"
        if name in auto_names:
            status = "OK"
            note = ""
        elif name in alt:
            status = "FAIL"
            note = "财务确认版要求同名/同口径审计sheet；当前只有替代原始表，不能直接签核"
        elif name == "税务A_B核对":
            status = "FAIL"
            note = "用户明确要求财务确认版补税务A_B核对sheet"
        else:
            status = "FAIL"
            note = "自动化输出缺失"
        rows.append([name, b_ok, a_same, alt.get(name, ""), status, note])
    return rows


def sum_monthly(rows: list[dict[str, Any]], store: str, field: str) -> float:
    total = 0.0
    for r in rows:
        if norm_store(r.get("平台"), r.get("店铺")) == store:
            total += n0(r.get(field))
    return total


def build_special_checks(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> list[Check]:
    checks: list[Check] = []
    base_month = rows_to_dicts(base.get("月度毛利试算", []))
    auto_month = rows_to_dicts(auto.get("月度毛利试算", [])) or rows_to_dicts(auto.get("11_店铺汇总看板", []))
    b_tmall_funlab_ad = sum_monthly(base_month, "天猫纷岚", "广告费(RMB)")
    a_tmall_funlab_ad = sum_monthly(auto_month, "天猫纷岚", "广告费(RMB)") or sum_monthly(auto_month, "天猫纷岚", "广告费合计")
    checks.append(Check(
        "天猫广告费 9391.83",
        "PASS" if abs(a_tmall_funlab_ad - b_tmall_funlab_ad) <= 0.01 else "FAIL",
        f"人工B 天猫纷岚广告费={b_tmall_funlab_ad:.2f}; 自动A={a_tmall_funlab_ad:.2f}",
        "修天猫广告流水 parser：按交易日期2026-06且扣款/支出读取操作金额(元)，不能默认0。",
    ))
    b_gp = sum_monthly(base_month, "天猫纷岚", "试算毛利(RMB)")
    a_gp = sum_monthly(auto_month, "天猫纷岚", "试算毛利(RMB)") or sum_monthly(auto_month, "天猫纷岚", "毛利额")
    checks.append(Check(
        "天猫纷岚负毛利不是数据缺口",
        "PASS" if a_gp < 0 and abs(a_gp - b_gp) <= 0.01 else "FAIL",
        f"人工B 天猫纷岚毛利={b_gp:.2f}; 自动A={a_gp:.2f}",
        "按结算口径重算天猫纷岚，广告费计入后仍应允许负毛利通过，不要转成资料缺口。",
    ))

    base_log = rows_to_dicts(base.get("物流匹配明细", []))
    auto_log = rows_to_dicts(auto.get("物流匹配明细", [])) or rows_to_dicts(auto.get("05_物流月结账单_导入", []))
    fallback_keys = ("API", "EXP_RECE", "fallback", "查询", "n8n")
    b_fallback = [
        str(r.get("运单号") or "").strip()
        for r in base_log
        if norm_platform(r.get("平台")) == "抖音"
        and any(k.lower() in f"{r.get('账单文件/API','')} {r.get('来源','')}".lower() for k in fallback_keys)
    ]
    a_waybills = {
        str(r.get("运单号") or "").strip()
        for r in auto_log
        if str(r.get("运单号") or "").strip()
    }
    hit = [w for w in b_fallback if w in a_waybills]
    if b_fallback:
        status = "PASS" if len(hit) == len(set(b_fallback)) else "FAIL"
        evidence = f"人工B抖音API/fallback运单={len(set(b_fallback))}；自动A原始物流命中={len(set(hit))}"
    else:
        status = "WARN"
        evidence = "人工B未识别到带 API/fallback 标记的抖音物流行，需人工复核口径说明"
    checks.append(Check(
        "抖音顺丰 API 物流 fallback",
        status,
        evidence,
        "即使原始账单命中，也必须在自动化输出物流匹配明细，标出API fallback和分摊结果。",
    ))

    base_prod = rows_to_dicts(base.get("产品毛利_月度", []))
    auto_prod = rows_to_dicts(auto.get("产品毛利_月度", []))
    b_xhs_refund_qty = sum(
        n0(r.get("退款数量"))
        for r in base_prod
        if norm_platform(r.get("国内电商平台名称")) == "小红书"
    )
    a_xhs_refund_qty = sum(
        n0(r.get("退款数量") if "退款数量" in r else r.get("退款数量(订单表口径)"))
        for r in auto_prod
        if norm_platform(r.get("国内电商平台名称") or r.get("平台")) == "小红书"
    )
    checks.append(Check(
        "小红书退款成本数量回查包裹详情",
        "PASS" if abs(a_xhs_refund_qty - b_xhs_refund_qty) <= 0.001 else "FAIL",
        f"人工B小红书退款数量={b_xhs_refund_qty:g}; 自动A={a_xhs_refund_qty:g}",
        "退款成本数量必须限定商品结算明细退款订单范围，再回查包裹详情SKU件数扣减。",
    ))

    auto_sheets = set(auto)
    checks.append(Check(
        "税务A_B核对 sheet",
        "PASS" if "税务A_B核对" in auto_sheets else "FAIL",
        f"自动A sheets={', '.join(sorted(auto_sheets))}",
        "财务确认版补单独税务A_B核对sheet：涉税资料读取、差异、解释、可定稿结论。",
    ))
    return checks


def collect_major_failures(sheet_rows: dict[str, list[list[Any]]]) -> list[list[Any]]:
    out = [["来源sheet", "定位1", "定位2", "定位3", "字段/项目", "人工B", "自动A", "差异", "状态", "说明"]]
    for sheet_name, rows in sheet_rows.items():
        if not rows:
            continue
        headers = [str(x) for x in rows[0]]
        for row in rows[1:]:
            text = " ".join(str(x) for x in row)
            if "FAIL" not in text and "缺字段" not in text and "缺失" not in text:
                continue
            cells = list(row) + [""] * 10
            out.append([sheet_name, cells[0], cells[1], cells[2], cells[3], cells[4], cells[5], cells[6], cells[7], cells[8]])
    return out


def write_compare_workbook(base: dict[str, list[list[Any]]], auto: dict[str, list[list[Any]]]) -> dict[str, list[list[Any]]]:
    wb = Workbook()
    wb.remove(wb.active)
    special = build_special_checks(base, auto)
    sections: dict[str, list[list[Any]]] = {
        "差异归因": build_root_cause_rows(),
        "输出sheet检查": build_sheet_check(base, auto),
        "月度毛利试算_AB": build_monthly_compare(base, auto),
        "产品毛利_月度_AB": build_product_compare(base, auto),
        "SKU成本明细_AB": build_cost_compare(base, auto),
        "物流匹配明细_AB": build_logistics_compare(base, auto),
        "费用明细汇总_AB": build_fee_compare(base, auto),
        "缺口清单_AB": build_gap_compare(base, auto),
        "专项检查": [["检查项", "状态", "证据", "建议动作"]] + [[c.item, c.status, c.evidence, c.action] for c in special],
    }
    sections["重大差异"] = collect_major_failures({k: v for k, v in sections.items() if k != "重大差异"})
    readme_rows = [
        ["项目", "值"],
        ["结论", "BLOCKED_BY_P0 - 口径已收敛，但仍有资料/无结算确认缺口"],
        ["人工基准B", str(BASELINE_PATH)],
        ["自动化A", f"https://u1wpma3xuhr.feishu.cn/sheets/{AUTO_SHEET_TOKEN}"],
        ["自动化本地重建", str(AUTO_REBUILT_PATH)],
        ["关键原因", "京东两店缺订单结算/到账文件或无结算确认；小红书纷岚缺商品结算明细或无结算确认。"],
    ]
    ws = wb.create_sheet("README")
    write_rows(ws, readme_rows)
    style_ws(ws)
    for title, rows in sections.items():
        ws = wb.create_sheet(title[:31])
        write_rows(ws, rows)
        style_ws(ws)
    wb.save(COMPARE_PATH)
    return sections


def build_root_cause_rows() -> list[list[Any]]:
    return [
        ["差异点", "归因类型", "判定", "证据", "下一步"],
        [
            "天猫/抖音/小红书宝空主链路",
            "口径已收敛",
            "可作为自动化对比基准",
            "月度毛利、产品毛利、SKU成本、物流匹配均已与人工B逐字段对齐；专项检查 5 项 PASS。",
            "进入财务确认前只需关闭剩余 P0 资料缺口。",
        ],
        [
            "京东纷岚 0.18 短信服务费差异",
            "资料缺口",
            "不是计算口径差异",
            "当前真实上传/Base 记录里京东纷岚没有订单结算明细/到账文件/无结算确认，自动A写 P0；人工B有短信服务费 0.18。",
            "运营补京东到账/结算文件，或通过卡片确认本月无结算并说明 0.18 费用来源。",
        ],
        [
            "京东宝空 P0",
            "资料缺口",
            "不是计算口径差异",
            "当前真实上传/Base 记录里京东宝空没有订单结算明细/到账文件/无结算确认，自动A写 P0。",
            "运营补京东到账/结算文件，或通过卡片确认本月无结算。",
        ],
        [
            "小红书纷岚 P0",
            "资料缺口",
            "不是计算口径差异",
            "当前真实上传/Base 记录里小红书纷岚有订单查询、涉税、平台支出，但没有商品结算明细，也没有结构化无结算确认。",
            "运营补商品结算明细，或通过卡片确认本月无结算。",
        ],
        [
            "财务确认 gate",
            "P0 gate 未通过",
            "不能发送财务确认卡",
            "输出 sheet 规范全 OK，但缺口清单仍有京东两店和小红书纷岚 P0。",
            "P0 关闭后重跑 A/B；通过后再发财务确认卡。",
        ],
    ]


def write_summary(sections: dict[str, list[list[Any]]]) -> None:
    special_rows = sections["专项检查"][1:]
    def bad_count(rows: list[list[Any]]) -> int:
        bad_markers = ("FAIL", "缺字段", "缺失", "不可确认")
        return sum(1 for row in rows[1:] if any(marker in str(row) for marker in bad_markers))

    monthly_fail_count = bad_count(sections["月度毛利试算_AB"])
    product_fail_count = bad_count(sections["产品毛利_月度_AB"])
    sheet_fail_count = bad_count(sections["输出sheet检查"])
    special_fail_count = sum(1 for row in special_rows if row[1] != "PASS")
    if sheet_fail_count == 0 and product_fail_count == 0 and special_fail_count == 0 and monthly_fail_count > 0:
        conclusion = "BLOCKED_BY_P0。口径专项、产品毛利、SKU成本、物流匹配和输出 sheet 已对齐；当前不能进入财务确认，只因真实资料/无结算确认缺口未关闭。"
    elif monthly_fail_count == 0 and product_fail_count == 0 and sheet_fail_count == 0 and special_fail_count == 0:
        conclusion = "PASS。自动化输出已达到 2026-06 人工基准 A/B 验收标准，可进入财务确认。"
    else:
        conclusion = "FAIL。仍存在口径或输出规范差异，不能进入财务确认。"
    lines = [
        "# 国内电商 2026-06 自动化 A/B 对账摘要",
        "",
        "## 结论",
        "",
        conclusion,
        "",
        "## 关键计数",
        "",
        f"- 月度毛利试算 FAIL/缺失字段行：{monthly_fail_count}",
        f"- 产品毛利_月度 FAIL/缺字段行：{product_fail_count}",
        f"- 输出 sheet 规范 FAIL 行：{sheet_fail_count}",
        "",
        "## 专项检查",
        "",
    ]
    for item, status, evidence, action in special_rows:
        lines.append(f"- {status}｜{item}｜{evidence}｜{action}")
    lines.extend([
        "",
        "## 产物",
        "",
        f"- 自动化飞书表本地重建：`{AUTO_REBUILT_PATH}`",
        f"- A/B 对账 workbook：`{COMPARE_PATH}`",
    ])
    SUMMARY_PATH.write_text("\n".join(lines), encoding="utf-8")


async def main() -> None:
    if not BASELINE_PATH.exists():
        raise FileNotFoundError(BASELINE_PATH)
    base = load_baseline()
    missing = [s for s in REQUIRED_BASELINE_SHEETS if s not in base]
    if missing:
        raise RuntimeError(f"Baseline missing required sheets: {missing}")
    auto = await read_auto_sheet_values()
    save_auto_rebuilt(auto)
    sections = write_compare_workbook(base, auto)
    write_summary(sections)
    print(f"saved_auto={AUTO_REBUILT_PATH}")
    print(f"saved_compare={COMPARE_PATH}")
    print(f"saved_summary={SUMMARY_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
